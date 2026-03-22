"""
pipeline/wide_irt.py
Procesa base bruta IRT → formato Wide (_IRT1/_IRT2/_IRT3).
Misma interfaz que wide_top.procesar_wide.
"""
import pandas as pd
import numpy as np
import unicodedata, re, warnings
from io import BytesIO
warnings.filterwarnings('ignore')

def _norm_str(s):
    return unicodedata.normalize('NFD', str(s).lower()).encode('ascii','ignore').decode()

MESES_ES = {1:'Enero',2:'Febrero',3:'Marzo',4:'Abril',5:'Mayo',6:'Junio',
            7:'Julio',8:'Agosto',9:'Septiembre',10:'Octubre',11:'Noviembre',12:'Diciembre'}

def auto_col(cols, keywords, nombre_col):
    for c in cols:
        if any(_norm_str(k) in _norm_str(c) for k in keywords):
            return c
    raise ValueError(f"Columna '{nombre_col}' no encontrada.")

def _parse_fecha(serie):
    if pd.api.types.is_datetime64_any_dtype(serie): return serie
    _MES = {'ene':'Jan','feb':'Feb','mar':'Mar','abr':'Apr','may':'May','jun':'Jun',
            'jul':'Jul','ago':'Aug','sept':'Sep','sep':'Sep','oct':'Oct','nov':'Nov','dic':'Dec'}
    def _conv(val):
        s = str(val).strip().lower()
        for es,en in _MES.items(): s = re.sub(rf'\b{es}\b', en, s)
        return pd.to_datetime(s, errors='coerce')
    result = pd.to_datetime(serie, errors='coerce')
    mask = result.isna() & serie.notna()
    if mask.any(): result[mask] = serie[mask].apply(_conv)
    return result

def norm_sust_v3(s):
    if pd.isna(s): return None
    raw = str(s).strip()
    if raw in ('0',''): return None
    raw = re.split(r'[\r\n]', raw)[0].strip()
    raw = re.sub(r'\(.*?\)','',raw).strip()
    raw = re.sub(r'^(las dos|ambas|los dos|ambos)[,\s]+','',raw,flags=re.IGNORECASE).strip()
    primera = re.split(r'\s+y\s+|[/,+]',raw,maxsplit=1)[0].strip()
    n = _norm_str(primera)
    if any(x in n for x in ['ninguno','ninguna','niega','no aplica','no consume','nada']): return None
    if any(x in n for x in ['ludopatia','juego','apuesta']): return None
    if any(x in n for x in ['alcohol','alchol','cerveza','licor','aguard','ron']): return 'Alcohol'
    if any(x in n for x in ['marihu','marhuana','cannabis','marij','weed','crispy']): return 'Marihuana'
    if any(x in n for x in ['tusi','tussi','tusy','tuci','2cb']): return 'Tusi'
    if any(x in n for x in ['pasta base','papelillo','pbc','basuco','bazuco']): return 'Pasta Base/Basuco'
    if any(x in n for x in ['metanfet','anfetam','cristal','crystal']): return 'Metanfetamina'
    if any(x in n for x in ['crack','piedra','paco']): return 'Crack'
    if any(x in n for x in ['cocain','cocai','perico','coke']): return 'Cocaína'
    if any(x in n for x in ['tabaco','cigarr','nicot']): return 'Tabaco'
    if any(x in n for x in ['inhalant','thiner','activo','pegamento']): return 'Inhalantes'
    if any(x in n for x in ['sedant','benzod','tranqui','clonaz','rivotril']): return 'Sedantes'
    if any(x in n for x in ['opiod','heroina','morfin','fentanil','tramad']): return 'Opiáceos'
    if any(x in n for x in ['extasis','mdma']): return 'Éxtasis'
    if any(x in n for x in ['ketam']): return 'Ketamina'
    return None

_SUST_KEYS = [_norm_str(x) for x in
    ['sustancia principal','cual considera','cuál considera','genera mas problemas']]


def procesar_wide(input_path, filtro_centro=None, fecha_desde=None, fecha_hasta=None):
    logs = []
    df = pd.read_excel(input_path, sheet_name=0, header=0)
    logs.append(f"✓ {len(df)} filas × {len(df.columns)} columnas")

    COL_CODIGO = auto_col(df.columns,
        ['identificacion','identificación','2 primeras letras','primer nombre','cod_pac'],
        'Código paciente')
    COL_FECHA = auto_col(df.columns,
        ['fecha de administracion','fecha_administracion','fecha administracion',
         'fecha entrevista','fecha_entrevista'],
        'Fecha de Administración')

    hoy = pd.Timestamp.now()
    df[COL_FECHA] = _parse_fecha(df[COL_FECHA])
    alertas = []

    COL_CENTRO = None
    for c in df.columns:
        nc = _norm_str(c)
        if any(k in nc for k in ['codigo del centro','centro de tratamiento','servicio de tratamiento']):
            if 'trabajo' not in nc and 'estudio' not in nc:
                COL_CENTRO = c; break

    if filtro_centro and COL_CENTRO:
        n_a = len(df)
        df = df[df[COL_CENTRO].astype(str).str.strip()==filtro_centro.strip()].copy().reset_index(drop=True)
        logs.append(f"✓ Filtro centro: {n_a}→{len(df)}")
        if len(df)==0: raise ValueError(f"Centro '{filtro_centro}' sin registros.")

    if fecha_desde or fecha_hasta:
        mask = pd.Series([True]*len(df), index=df.index)
        if fecha_desde: mask &= df[COL_FECHA] >= pd.Timestamp(fecha_desde+'-01')
        if fecha_hasta: mask &= df[COL_FECHA] <= pd.Timestamp(fecha_hasta+'-01') + pd.offsets.MonthEnd(0)
        n_a = len(df); df = df[mask].copy().reset_index(drop=True)
        logs.append(f"✓ Filtro período: {n_a}→{len(df)}")
        if len(df)==0: raise ValueError("Sin registros en el período.")

    centro_lookup = df.groupby(COL_CODIGO)[COL_CENTRO].first().to_dict() if COL_CENTRO else {}
    def get_centro(cod): return str(centro_lookup.get(cod,'—'))[:60]

    COL_FN = next((c for c in df.columns if c != COL_CODIGO and
                   any(k in _norm_str(c) for k in ['fecha de nacimiento','fecha_nacimiento'])), None)
    if COL_FN:
        df[COL_FN] = _parse_fecha(df[COL_FN])
        for idx,row in df.iterrows():
            fn=row[COL_FN]; cod=row[COL_CODIGO]
            if pd.isna(fn): continue
            if fn>hoy:
                alertas.append({'Código':cod,'Centro':get_centro(cod),'Columna':COL_FN,'Valor':str(fn.date()),'Regla':'Fecha nacimiento futura'})
                df.at[idx,COL_FN]=np.nan; continue
            edad=(hoy-fn).days/365.25
            if edad<10 or edad>100:
                alertas.append({'Código':cod,'Centro':get_centro(cod),'Columna':COL_FN,'Valor':str(fn.date()),'Regla':f'Edad={edad:.1f} años'})
                df.at[idx,COL_FN]=np.nan

    for c in [c for c in df.columns if '(0-7)' in c and 'Promedio' not in c]:
        num=pd.to_numeric(df[c],errors='coerce'); mask_=num>7
        for idx in df[mask_].index:
            alertas.append({'Código':df.at[idx,COL_CODIGO],'Centro':get_centro(df.at[idx,COL_CODIGO]),'Columna':c,'Valor':df.at[idx,c],'Regla':'Días sem>7'})
            df.at[idx,c]=np.nan

    for c in [c for c in df.columns if 'Total (0-28)' in c and 'Promedio' not in c]:
        num=pd.to_numeric(df[c],errors='coerce')
        for idx in df[(num>28)|(num<0)].index:
            alertas.append({'Código':df.at[idx,COL_CODIGO],'Centro':get_centro(df.at[idx,COL_CODIGO]),'Columna':c,'Valor':df.at[idx,c],'Regla':'Días mes fuera 0-28'})
            df.at[idx,c]=np.nan

    logs.append(f"✓ {len(alertas)} valores corregidos")

    fechas_ok = df[COL_FECHA].dropna()
    fechas_ok = fechas_ok[(fechas_ok.dt.year>=hoy.year-10)&(fechas_ok.dt.year<=hoy.year+1)]
    if len(fechas_ok):
        f0,f1=fechas_ok.min(),fechas_ok.max()
        if f0.year==f1.year and f0.month==f1.month: periodo=f'{MESES_ES[f0.month]} {f0.year}'
        elif f0.year==f1.year: periodo=f'{MESES_ES[f0.month]}–{MESES_ES[f1.month]} {f0.year}'
        else: periodo=f'{MESES_ES[f0.month]} {f0.year} – {MESES_ES[f1.month]} {f1.year}'
    else: periodo='Período no determinado'

    df = df.sort_values([COL_CODIGO,COL_FECHA]).reset_index(drop=True)
    conteo  = df[COL_CODIGO].value_counts()
    N_total = int(conteo.shape[0])
    N_irt2  = int((conteo>=2).sum())
    N_irt3  = int((conteo>=3).sum())
    N_solo1 = N_total - N_irt2

    rows1,rows2,rows3=[],[],[]
    for cod,grp in df.groupby(COL_CODIGO,sort=False):
        grp=grp.reset_index(drop=True)
        rows1.append(grp.loc[0])
        if len(grp)>=2: rows2.append(grp.loc[1])
        if len(grp)>=3: rows3.append(grp.loc[2])

    df1=pd.DataFrame(rows1).reset_index(drop=True)
    otras=[c for c in df1.columns if c!=COL_CODIGO]
    t1=df1.rename(columns={c:f'{c}_IRT1' for c in otras})
    wide=t1.copy()

    if rows2:
        df2=pd.DataFrame(rows2).reset_index(drop=True)
        df2a=df2.set_index(COL_CODIGO).reindex(df1[COL_CODIGO]).reset_index()
        t2=df2a.rename(columns={c:f'{c}_IRT2' for c in otras})
        wide=wide.merge(t2,on=COL_CODIGO,how='left')
    if rows3:
        df3=pd.DataFrame(rows3).reset_index(drop=True)
        df3a=df3.set_index(COL_CODIGO).reindex(df1[COL_CODIGO]).reset_index()
        t3=df3a.rename(columns={c:f'{c}_IRT3' for c in otras})
        wide=wide.merge(t3,on=COL_CODIGO,how='left')

    wide.insert(1,'Tiene_IRT1','Sí')
    irt2_cols=[c for c in wide.columns if c.endswith('_IRT2')]
    wide.insert(2,'Tiene_IRT2',
        wide[irt2_cols].notna().any(axis=1).map({True:'Sí',False:'No'}) if irt2_cols else 'No')

    _col_f1=next((c for c in wide.columns if 'fecha' in _norm_str(c) and c.endswith('_IRT1')),None)
    _HOY=pd.Timestamp.now().normalize()
    _n_rojo=_n_naranja=_n_verde=0
    if _col_f1:
        _f=pd.to_datetime(wide[_col_f1],errors='coerce')
        _d=(_HOY-_f).dt.days
        def _al(d):
            if pd.isna(d): return ''
            if d<60: return '🟢 <60 dias'
            if d<90: return '🟠 60-89 dias'
            return '🔴 90+ dias'
        wide['Dias_desde_IRT1']=_d.where(wide['Tiene_IRT2']=='No',other=None)
        wide['Alerta_IRT2']=_d.where(wide['Tiene_IRT2']=='No').apply(lambda d:_al(d) if not pd.isna(d) else '')
        wide.loc[wide['Tiene_IRT2']=='Sí','Alerta_IRT2']='Completado'
        _n_rojo=int((wide['Alerta_IRT2']=='🔴 90+ dias').sum())
        _n_naranja=int((wide['Alerta_IRT2']=='🟠 60-89 dias').sum())
        _n_verde=int((wide['Alerta_IRT2']=='🟢 <60 dias').sum())
    else:
        wide['Dias_desde_IRT1']=None; wide['Alerta_IRT2']=''

    for _sfx in ('_IRT1','_IRT2','_IRT3'):
        _col=next((c for c in wide.columns if c.endswith(_sfx) and
                   any(k in _norm_str(c) for k in _SUST_KEYS) and 'RAW' not in c),None)
        if not _col: continue
        _raw=_col.replace(_sfx,f'_RAW{_sfx}')
        wide.rename(columns={_col:_raw},inplace=True)
        wide[_col]=wide[_raw].apply(norm_sust_v3)
        _i=wide.columns.get_loc(_raw)
        wide=wide[[*wide.columns[:_i+1],_col,*[c for c in wide.columns[_i+1:] if c!=_col]]]
        logs.append(f"✓ Sustancia normalizada {_sfx}")

    dupes_data=[]
    for _,row in df[df.duplicated([COL_CODIGO,COL_FECHA],keep=False)][[COL_CODIGO,COL_FECHA]].drop_duplicates().iterrows():
        dupes_data.append({'Código':row[COL_CODIGO],'Fecha':str(row[COL_FECHA])[:10]})

    excel_bytes=_excel_wide(wide,alertas,dupes_data,COL_CODIGO,COL_CENTRO,_col_f1,
                            N_total,N_irt2,N_irt3,N_solo1,len(alertas),len(dupes_data),
                            _n_rojo,_n_naranja,_n_verde,periodo)

    col_sust=next((c for c in wide.columns if any(k in _norm_str(c) for k in _SUST_KEYS)
                   and c.endswith('_IRT1') and 'RAW' not in c),None)
    sust_dist=wide[col_sust].dropna().value_counts().head(8).to_dict() if col_sust else {}

    centros=[]
    if COL_CENTRO:
        ccw=f'{COL_CENTRO}_IRT1'
        if ccw in wide.columns:
            apps=df.groupby(COL_CENTRO).size().reset_index(name='Aplicaciones').rename(columns={COL_CENTRO:'Centro'})
            res=wide.groupby(ccw).agg(Pacientes=(COL_CODIGO,'count'),Con_IRT2=('Tiene_IRT2',lambda x:(x=='Sí').sum())).reset_index().rename(columns={ccw:'Centro'})
            res['Sin_IRT2']=res['Pacientes']-res['Con_IRT2']
            if alertas:
                df_al=pd.DataFrame(alertas)
                corr=df_al.groupby('Centro').size().reset_index(name='Vals_corregidos')
                res=res.merge(corr,on='Centro',how='left')
            else: res['Vals_corregidos']=0
            res['Vals_corregidos']=res['Vals_corregidos'].fillna(0).astype(int)
            res=res.merge(apps,on='Centro',how='left')
            res['Aplicaciones']=res['Aplicaciones'].fillna(0).astype(int)
            res=res.sort_values('Aplicaciones',ascending=False)
            tots={'Centro':'TOTAL','Aplicaciones':int(res['Aplicaciones'].sum()),
                  'Pacientes':int(res['Pacientes'].sum()),'Con_IRT2':int(res['Con_IRT2'].sum()),
                  'Sin_IRT2':int(res['Sin_IRT2'].sum()),'Vals_corregidos':int(res['Vals_corregidos'].sum())}
            centros=res[['Centro','Aplicaciones','Pacientes','Con_IRT2','Sin_IRT2','Vals_corregidos']].to_dict('records')
            centros.append(tots)

    return {'wide':wide,'filtro_centro':filtro_centro,'fecha_desde':fecha_desde,'fecha_hasta':fecha_hasta,
            'stats':{'N_total':N_total,'N_irt2':N_irt2,'N_irt3':N_irt3,'N_solo1':N_solo1,
                     'pct_irt2':round(N_irt2/N_total*100,1) if N_total else 0,
                     'N_alertas':len(alertas),'N_dupes':len(dupes_data),
                     'n_rojo':_n_rojo,'n_naranja':_n_naranja,'n_verde':_n_verde,
                     'cols_wide':len(wide.columns),'sust_dist':sust_dist},
            'centros':centros,'alertas':alertas,'dupes':dupes_data,
            'periodo':periodo,'excel_bytes':excel_bytes,'logs':logs}


def _excel_wide(wide,alertas,dupes,COL_CODIGO,COL_CENTRO,col_f1,
                N_total,N_irt2,N_irt3,N_solo1,N_al,N_du,
                n_rojo,n_naranja,n_verde,periodo):
    from openpyxl import Workbook
    from openpyxl.styles import Font,PatternFill,Alignment
    C_DARK='1F3864'; C_WHITE='FFFFFF'; C_ALT='EEF4FB'
    wb=Workbook(); ws=wb.active; ws.title='Base Wide'
    ws.sheet_view.showGridLines=False; ws.freeze_panes='B3'
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=len(wide.columns))
    ct=ws.cell(1,1); ct.value=f'QALAT · Base Wide IRT · {periodo} · {N_total} pacientes'
    ct.font=Font(bold=True,size=10,color=C_WHITE,name='Arial')
    ct.fill=PatternFill('solid',start_color=C_DARK)
    ct.alignment=Alignment(horizontal='center',vertical='center')
    ws.row_dimensions[1].height=22
    for ci,col in enumerate(wide.columns,1):
        c=ws.cell(2,ci); c.value=col
        c.font=Font(bold=True,size=8,color=C_WHITE,name='Arial')
        c.fill=PatternFill('solid',start_color=C_DARK)
        c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
        ws.column_dimensions[c.column_letter].width=max(10,min(35,len(str(col))*0.85))
    ws.row_dimensions[2].height=40
    for ri,row in wide.iterrows():
        bg=PatternFill('solid',start_color=(C_ALT if ri%2==0 else C_WHITE))
        for ci,val in enumerate(row,1):
            c=ws.cell(ri+3,ci)
            c.value=None if (not isinstance(val,str) and pd.isna(val)) else val
            c.font=Font(size=8,name='Arial'); c.fill=bg
            c.alignment=Alignment(horizontal='center',vertical='center')
    wr=wb.create_sheet('Resumen'); wr.sheet_view.showGridLines=False
    for ri,(k,v) in enumerate([('Instrumento','IRT'),('Período',periodo),
            ('Pacientes únicos',N_total),('Con IRT2',N_irt2),('Con IRT3',N_irt3),
            ('Solo IRT1',N_solo1),('% con seguimiento',f'{round(N_irt2/N_total*100,1) if N_total else 0}%'),
            ('Valores corregidos',N_al),('🔴 Urgentes',n_rojo),
            ('🟠 Próximos',n_naranja),('🟢 Con tiempo',n_verde)],1):
        wr.cell(ri,1).value=k; wr.cell(ri,1).font=Font(bold=True,size=9,name='Arial')
        wr.cell(ri,2).value=v; wr.cell(ri,2).font=Font(size=9,name='Arial')
    wr.column_dimensions['A'].width=28; wr.column_dimensions['B'].width=20
    wa=wb.create_sheet('Alertas'); wa.sheet_view.showGridLines=False
    for ci,h in enumerate(['Código','Centro','Columna','Valor','Regla'],1):
        c=wa.cell(1,ci); c.value=h
        c.font=Font(bold=True,size=8,color=C_WHITE,name='Arial')
        c.fill=PatternFill('solid',start_color=C_DARK)
    for ri,a in enumerate(alertas,2):
        for ci,k in enumerate(['Código','Centro','Columna','Valor','Regla'],1):
            wa.cell(ri,ci).value=a.get(k,''); wa.cell(ri,ci).font=Font(size=8,name='Arial')
    wq=wb.create_sheet('Calidad de Datos'); wq.sheet_view.showGridLines=False
    if dupes:
        for ci,h in enumerate(['Código','Fecha'],1):
            c=wq.cell(1,ci); c.value=h
            c.font=Font(bold=True,size=8,color=C_WHITE,name='Arial')
            c.fill=PatternFill('solid',start_color=C_DARK)
        for ri,d in enumerate(dupes,2):
            wq.cell(ri,1).value=d['Código']; wq.cell(ri,2).value=d['Fecha']
    else:
        wq.cell(1,1).value='✅ Sin fechas duplicadas'
    buf=BytesIO(); wb.save(buf); buf.seek(0)
    return buf
