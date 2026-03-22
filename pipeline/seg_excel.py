"""
╔══════════════════════════════════════════════════════════════════════════════╗
║   SCRIPT_IRT_Universal_Seguimiento_Excel.py  —  v1.1                      ║
║   Genera tablas de seguimiento IRT1 vs IRT2 (vs IRT3 si existe)           ║
║   Compatible con cualquier país que use el instrumento IRT                 ║
╠══════════════════════════════════════════════════════════════════════════════╣
║  CÓMO USAR:                                                                 ║
║  1. Abre un chat nuevo con Claude                                           ║
║  2. Sube DOS archivos:                                                      ║
║       • Este script                                                         ║
║       • La base Wide IRT (generada por SCRIPT_IRT_Universal_Wide)          ║
║  3. Escribe: "Ejecuta el script Seguimiento IRT con esta base Wide"        ║
║                                                                             ║
║  TABLAS GENERADAS (IRT1 vs IRT2, y vs IRT3 si existe):                    ║
║    0.   Instrumentos Contestados (n por etapa)                             ║
║    1.1  Distribución por Sexo                                               ║
║    1.2  Distribución por Rango de Edad                                      ║
║    2.1  Sustancia Principal (IRT1 vs IRT2)                                 ║
║    2.2  Promedio Días Consumo — Sustancia Principal                        ║
║    2.3  Evolución del Consumo — Sustancia Principal                        ║
║    2.4  Cambio de Consumo por Sustancia (IRT1 vs IRT2)                    ║
║    2.5  % Consumidores por Sustancia (IRT1 vs IRT2)                       ║
║    2.6  Promedio Días de Consumo por Sustancia (IRT1 vs IRT2)             ║
║    3.1  Urgencia u Hospitalización (IRT1 vs IRT2)                         ║
║    3.2  Accidentes por Consumo (IRT1 vs IRT2)                             ║
║    3.3  Autopercepción Salud Psicológica y Física (IRT1 vs IRT2)          ║
║    3.4  Problemas en Trabajo / Educación (IRT1 vs IRT2)                   ║
║    4.1  Transgresión — Total (IRT1 vs IRT2)                               ║
║    4.2  Transgresión — Por Tipo (IRT1 vs IRT2)                            ║
║    5.1  Relaciones Interpersonales (IRT1 vs IRT2)                         ║
║    6.1  Satisfacción de Vida (IRT1 vs IRT2)                               ║
╚══════════════════════════════════════════════════════════════════════════════╝
"""
import glob, os, unicodedata, warnings
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
warnings.filterwarnings('ignore')

# ── Colores ───────────────────────────────────────────────────────────────────
C_DARK  = '1F3864'; C_MID   = '2E75B6'; C_LIGHT = 'BDD7EE'
C_ALT   = 'EEF4FB'; C_WHITE = 'FFFFFF'; C_BDR   = 'B8CCE4'
C_NOTE  = '595959'; C_IRT2  = '00B0F0'; C_IRT3  = '70AD47'
C_BAJA  = 'C00000'; C_SUBE  = '375623'; C_IGUAL = '595959'

def _norm(s):
    return unicodedata.normalize('NFD', str(s).lower()).encode('ascii','ignore').decode()

# ── Auto-detección de archivo Wide ───────────────────────────────────────────
def auto_archivo_wide():
    candidatos = (
        glob.glob('/mnt/user-data/uploads/*IRT*Wide*.xlsx') +
        glob.glob('/mnt/user-data/uploads/*Wide*IRT*.xlsx') +
        glob.glob('/mnt/user-data/uploads/IRT_Base*.xlsx') +
        glob.glob('/mnt/user-data/outputs/IRT_Base_Wide*.xlsx') +
        glob.glob('/home/claude/IRT_Base_Wide.xlsx'))
    if not candidatos:
        raise FileNotFoundError(
            "\n⚠  No se encontró la base Wide IRT.\n"
            "   Sube el archivo IRT_Base_Wide.xlsx junto con este script.")
    uploads = [f for f in candidatos if 'uploads' in f]
    elegido = uploads[0] if uploads else max(candidatos, key=os.path.getsize)
    print(f"  → Base Wide detectada: {os.path.basename(elegido)}")
    return elegido

# ══════════════════════════════════════════════════════════════════════════════
print('=' * 60)
print('  SCRIPT_IRT_Universal_Seguimiento_Excel  v1.1')
print('=' * 60)

INPUT_FILE  = auto_archivo_wide()

# ── FILTRO OPCIONAL POR CENTRO ────────────────────────────────────────────────
# Dejar en None para procesar TODOS los centros.
# Poner el código exacto del centro para filtrar solo ese centro.
# Ejemplos:
#   FILTRO_CENTRO = None         ← todos los centros
#   FILTRO_CENTRO = "HCHN01"     ← solo ese centro
FILTRO_CENTRO = None
# ─────────────────────────────────────────────────────────────────────────────

OUTPUT_FILE = '/home/claude/IRT_Seguimiento.xlsx'

df = pd.read_excel(INPUT_FILE, sheet_name='Base Wide', header=1)
df.columns = [str(c) for c in df.columns]
cols = df.columns.tolist()

# Aplicar filtro de centro si corresponde
_col_centro = next((c for c in cols if any(x in _norm(c) for x in
                    ['codigo del centro', 'servicio de tratamiento',
                     'centro/ servicio', 'codigo centro'])), None)
if FILTRO_CENTRO and _col_centro:
    n_antes = len(df)
    df = df[df[_col_centro].astype(str).str.strip() == FILTRO_CENTRO].copy()
    df = df.reset_index(drop=True)
    print(f'\n⚑  Filtro activo: Centro = "{FILTRO_CENTRO}"')
    print(f'   {n_antes} pacientes totales → {len(df)} del centro seleccionado')
    OUTPUT_FILE = f'/home/claude/IRT_Seguimiento_{FILTRO_CENTRO}.xlsx'
elif FILTRO_CENTRO and not _col_centro:
    print(f'\n⚠  FILTRO_CENTRO = "{FILTRO_CENTRO}" pero no se encontró columna de centro.')

N_total = len(df)
mask2   = df['Tiene_IRT2'] == 'Sí'
mask3   = df['Tiene_IRT3'] == 'Sí'
N_irt2  = int(mask2.sum())
N_irt3  = int(mask3.sum())
TIENE_IRT3 = N_irt3 > 0

print(f'\n→ {N_total} pacientes cargados')
print(f'  IRT2 (seguimiento 3m): {N_irt2} | IRT3 (seguimiento 6m): {N_irt3}')

if N_irt2 == 0:
    print('\n⚠  No hay pacientes con IRT2. No es posible generar el informe de seguimiento.')
    raise SystemExit(0)

# ── Detectar período y servicio ───────────────────────────────────────────────
MESES_ES = {1:'Enero',2:'Febrero',3:'Marzo',4:'Abril',5:'Mayo',6:'Junio',
            7:'Julio',8:'Agosto',9:'Septiembre',10:'Octubre',
            11:'Noviembre',12:'Diciembre'}

fecha_col = next((c for c in cols if 'fecha de administracion' in _norm(c)
                  or 'fecha_aplicacion' in _norm(c)), None)
if fecha_col:
    fechas = pd.to_datetime(df[fecha_col], errors='coerce').dropna()
    anio   = pd.Timestamp.now().year
    fechas = fechas[(fechas.dt.year >= anio-10) & (fechas.dt.year <= anio+1)]
    if len(fechas):
        f0, f1 = fechas.min(), fechas.max()
        if f0.year == f1.year and f0.month == f1.month:
            PERIODO = f'{MESES_ES[f0.month]} {f0.year}'
        elif f0.year == f1.year:
            PERIODO = f'{MESES_ES[f0.month]}–{MESES_ES[f1.month]} {f0.year}'
        else:
            PERIODO = f'{MESES_ES[f0.month]} {f0.year} – {MESES_ES[f1.month]} {f1.year}'
    else:
        PERIODO = 'Período no determinado'
else:
    PERIODO = 'Período no determinado'

if _col_centro:
    centros = df[_col_centro].dropna().unique()
    SERVICIO = centros[0] if len(centros) == 1 else f'{len(centros)} centros'
else:
    SERVICIO = 'Servicio de Tratamiento'

print(f'  Período : {PERIODO}')
print(f'  Servicio: {SERVICIO}')

# ══════════════════════════════════════════════════════════════════════════════
# DETECTAR COLUMNAS CLAVE
# ══════════════════════════════════════════════════════════════════════════════
def col_sfx(keywords, sfx, obligatoria=False):
    for c in cols:
        if not c.endswith(sfx): continue
        if all(_norm(k) in _norm(c) for k in keywords):
            return c
    return None

COL_FN   = next((c for c in cols if 'fecha de nacimiento' in _norm(c)), None)
COL_SEXO = next((c for c in cols if _norm(c) in ['sexo','género','genero']), None)

# Sustancias
SUST_NOMBRES = {
    'Alcohol':        ['alcohol'],
    'Marihuana':      ['marihuana','cannabis'],
    'Heroína':        ['heroina','heroin'],
    'Cocaína':        ['cocain'],
    'Fentanilo':      ['fentanil'],
    'Inhalables':     ['inhalab'],
    'Metanfetamina':  ['metanfet','cristal'],
    'Crack':          ['crack'],
    'Pasta Base':     ['pasta base','pasta'],
    'Sedantes':       ['sedant','benzod'],
    'Opiáceos':       ['opiod','opiac'],
    'Tabaco':         ['tabaco','nicot'],
    'Otra sustancia': ['otra sust'],
}

SUST_TOTAL = {}  # sust → {IRT1: col, IRT2: col, IRT3: col}
for sust, kws in SUST_NOMBRES.items():
    entry = {}
    for sfx in ['_IRT1','_IRT2','_IRT3']:
        for c in cols:
            if not c.endswith(sfx): continue
            nc = _norm(c)
            if not any(_norm(k) in nc for k in kws): continue
            if 'total' in nc or '(0-28)' in nc:
                entry[sfx] = c; break
    if entry:
        SUST_TOTAL[sust] = entry
SUST_ACTIVAS = list(SUST_TOTAL.keys())

COL_SP = {sfx: col_sfx(['sustancia','principal'], sfx) for sfx in ['_IRT1','_IRT2','_IRT3']}

# Salud
COL_SPSI = {sfx: col_sfx(['salud','psicol'], sfx) for sfx in ['_IRT1','_IRT2','_IRT3']}
COL_SFIS = {sfx: col_sfx(['salud','fis'],    sfx) for sfx in ['_IRT1','_IRT2','_IRT3']}

# Urgencias / Accidentes
COL_URG  = {sfx: next((c for c in cols if c.endswith(sfx) and '5)' in c
                        and any(k in c.lower() for k in ['urgencia','hospitali','emergencia'])), None)
            for sfx in ['_IRT1','_IRT2','_IRT3']}
COL_ACC  = {sfx: next((c for c in cols if c.endswith(sfx) and '6)' in c
                        and 'accidente' in c.lower()), None)
            for sfx in ['_IRT1','_IRT2','_IRT3']}

# Trabajo
COL_TARDE    = {sfx: next((c for c in cols if c.endswith(sfx) and '10)' in c and
                            any(k in c.lower() for k in ['tarde','atrasado','antes de su jornada'])), None)
               for sfx in ['_IRT1','_IRT2','_IRT3']}
COL_FALTAS   = {sfx: next((c for c in cols if c.endswith(sfx) and '10)' in c and 'falt' in c.lower()), None)
               for sfx in ['_IRT1','_IRT2','_IRT3']}
COL_SANCIONES= {sfx: next((c for c in cols if c.endswith(sfx) and '10)' in c and 'sancion' in c.lower()), None)
               for sfx in ['_IRT1','_IRT2','_IRT3']}

# Transgresión
TRANS_COLS = {}
for nombre, kw in [('Robo / Hurto','robo'), ('Venta de sustancias','venta'),
                   ('Violencia a otras personas','violencia'),
                   ('Violencia intrafamiliar','intraf'), ('Detenido / Arrestado','detenido')]:
    entry = {}
    for sfx in ['_IRT1','_IRT2','_IRT3']:
        c = next((col for col in cols if col.endswith(sfx) and kw in col.lower()), None)
        if c: entry[sfx] = c
    if entry: TRANS_COLS[nombre] = entry

# Relaciones
REL_MAP = {'Padre':['padre'], 'Madre':['madre'], 'Hijos':['hijos','hijo'],
           'Hermanos':['hermanos'], 'Pareja':['pareja'],
           'Amigos':['amigos'], 'Otros':['otros']}
REL_COLS = {}
for vinculo, kws in REL_MAP.items():
    entry = {}
    for sfx in ['_IRT1','_IRT2','_IRT3']:
        for c in cols:
            if not c.endswith(sfx): continue
            nc = _norm(c)
            if '14)' not in c and 'relaci' not in nc: continue
            if any(_norm(k) in nc for k in kws):
                entry[sfx] = c; break
    if entry: REL_COLS[vinculo] = entry

# Satisfacción
SAT_MAP = {
    'Vida en general':     [['16)'], ['satisfac','vida']],
    'Lugar donde vive':    [['17)'], ['satisfac','lugar']],
    'Situación laboral':   [['18)'], ['satisfac','labor','educac']],
    'Tiempo libre':        [['19)'], ['satisfac','tiempo']],
    'Capacidad económica': [['20)'], ['satisfac','econom']],
}
SAT_COLS = {}
for dim, (nums, kws) in SAT_MAP.items():
    entry = {}
    for sfx in ['_IRT1','_IRT2','_IRT3']:
        for c in cols:
            if not c.endswith(sfx): continue
            nc = _norm(c)
            if any(n in c for n in nums) and any(_norm(k) in nc for k in kws):
                entry[sfx] = c; break
    if entry: SAT_COLS[dim] = entry

print(f'\n→ Detección:')
print(f'  Sustancias: {SUST_ACTIVAS}')
print(f'  Urgencias IRT2: {COL_URG["_IRT2"] is not None} | Accidentes IRT2: {COL_ACC["_IRT2"] is not None}')
print(f'  Relaciones: {list(REL_COLS.keys())}')
print(f'  Satisfacción: {list(SAT_COLS.keys())}')

# ══════════════════════════════════════════════════════════════════════════════
# HELPERS DE NORMALIZACIÓN
# ══════════════════════════════════════════════════════════════════════════════
def norm_sust(s):
    if pd.isna(s) or str(s).strip() in ['0','']: return None
    s = _norm(str(s))
    if any(x in s for x in ['alcohol','cerveza','licor','aguard']): return 'Alcohol'
    if any(x in s for x in ['marihu','cannabis','marij']):          return 'Marihuana'
    if any(x in s for x in ['crack','piedra','paco']):              return 'Crack'
    if any(x in s for x in ['pasta base','pasta']):                 return 'Pasta Base'
    if any(x in s for x in ['cocain','perico']):                    return 'Cocaína'
    if any(x in s for x in ['fentanil']):                           return 'Fentanilo'
    if any(x in s for x in ['inhalab','thiner']):                   return 'Inhalables'
    if any(x in s for x in ['metanfet','cristal','anfetam']):       return 'Metanfetamina'
    if any(x in s for x in ['sedant','benzod','valium']):           return 'Sedantes'
    if any(x in s for x in ['heroina','heroína','opiod']):          return 'Heroína'
    if any(x in s for x in ['tabaco','cigarr','nicot']):            return 'Tabaco'
    return 'Otra sustancia'

def calc_sino(col, mask):
    if col is None: return None, None, None
    vals = df.loc[mask, col].dropna().astype(str).str.strip().str.lower()
    nv   = len(vals)
    nsi  = int(vals.isin(['sí','si','yes','1','true']).sum())
    return nsi, round(nsi/nv*100, 1) if nv else 0.0, nv

def prom_col(col, mask):
    if col is None: return np.nan, 0
    vals = pd.to_numeric(df.loc[mask, col], errors='coerce').dropna()
    return (round(float(vals.mean()), 1) if len(vals) else np.nan), len(vals)

def n_trans(col, mask):
    if col is None: return 0, 0.0, 0
    vals = pd.to_numeric(df.loc[mask, col], errors='coerce').fillna(0)
    nv   = len(vals)
    nsi  = int((vals > 0).sum())
    return nsi, round(nsi/nv*100,1) if nv else 0.0, nv

# ══════════════════════════════════════════════════════════════════════════════
# CÁLCULOS
# ══════════════════════════════════════════════════════════════════════════════
print('\n→ Calculando tablas...')
hoy = pd.Timestamp.now()

# ── Sociodemografía (sobre pacientes con IRT2) ────────────────────────────────
if COL_SEXO:
    sexo = df.loc[mask2, COL_SEXO].value_counts(dropna=True)
    T11  = pd.DataFrame({'Categoría': sexo.index,
                         'n': sexo.values,
                         '%': (sexo.values/N_irt2*100).round(1)})
else:
    T11 = None

if COL_FN:
    edades = ((hoy - pd.to_datetime(df.loc[mask2, COL_FN], errors='coerce')).dt.days / 365.25).dropna()
    bins   = [0,17,25,35,45,55,200]
    labels = ['< 18 años','18–25 años','26–35 años','36–45 años','46–55 años','56 años o más']
    grupos = pd.cut(edades, bins=bins, labels=labels).value_counts().reindex(labels, fill_value=0)
    T12    = pd.DataFrame({'Rango': labels, 'n': grupos.values,
                           '%': (grupos.values/N_irt2*100).round(1)})
    edad_m, edad_min, edad_max = edades.mean(), edades.min(), edades.max()
else:
    T12 = None; edad_m = edad_min = edad_max = None

# ── 2.1 Sustancia principal IRT1 vs IRT2 ─────────────────────────────────────
def dist_sust_ppal(sfx, mask):
    col = COL_SP.get(sfx)
    if col is None: return {}
    cnt = df.loc[mask, col].apply(norm_sust).dropna().value_counts()
    return cnt.to_dict()

sp1 = dist_sust_ppal('_IRT1', mask2)
sp2 = dist_sust_ppal('_IRT2', mask2)
sp3 = dist_sust_ppal('_IRT3', mask3) if TIENE_IRT3 else {}
all_sust_pp = sorted(set(list(sp1)+list(sp2)+list(sp3)))
T21_rows = []
for s in all_sust_pp:
    r = {'Sustancia': s,
         'n IRT1': sp1.get(s,0), '% IRT1': round(sp1.get(s,0)/N_irt2*100,1),
         'n IRT2': sp2.get(s,0), '% IRT2': round(sp2.get(s,0)/N_irt2*100,1)}
    if TIENE_IRT3:
        r.update({'n IRT3': sp3.get(s,0), '% IRT3': round(sp3.get(s,0)/N_irt3*100,1)})
    T21_rows.append(r)
T21 = pd.DataFrame(T21_rows).sort_values('n IRT1', ascending=False)

# ── 2.2 Promedio días sustancia principal ────────────────────────────────────
T22_rows = []
for sust in SUST_ACTIVAS:
    entry = SUST_TOTAL[sust]
    # Solo personas cuya sust principal ES esta sust
    for sfx, msk, N_sfx in [('_IRT1', mask2, N_irt2),
                              ('_IRT2', mask2, N_irt2),
                              ('_IRT3', mask3, N_irt3)]:
        if sfx == '_IRT3' and not TIENE_IRT3: continue
        col_sp = COL_SP.get(sfx)
        if col_sp is None: continue
        mask_sp = msk & (df[col_sp].apply(norm_sust) == sust)
        col_tot = entry.get(sfx)
        if col_tot is None: continue
        vals = pd.to_numeric(df.loc[mask_sp, col_tot], errors='coerce').dropna()
        if len(vals) == 0: continue
        T22_rows.append({'Sustancia': sust, 'Etapa': sfx.replace('_',''),
                         'Promedio días (0–28)': round(vals.mean(),1), 'n': len(vals)})
T22 = pd.DataFrame(T22_rows) if T22_rows else None

# ── 2.3 Evolución consumo sustancia principal (redujo/no consumo/aumentó) ───
evol_rows = []
for sust in SUST_ACTIVAS:
    entry = SUST_TOTAL[sust]
    c1 = entry.get('_IRT1'); c2 = entry.get('_IRT2')
    if not c1 or not c2: continue
    # Solo quienes tenían esta sust como principal en IRT1
    col_sp1 = COL_SP.get('_IRT1')
    if col_sp1 is None: continue
    mask_pp = mask2 & (df[col_sp1].apply(norm_sust) == sust)
    n_pp = int(mask_pp.sum())
    if n_pp == 0: continue
    v1 = pd.to_numeric(df.loc[mask_pp, c1], errors='coerce')
    v2 = pd.to_numeric(df.loc[mask_pp, c2], errors='coerce')
    validos = v1.notna() & v2.notna()
    n_val   = int(validos.sum())
    if n_val == 0: continue
    diff = v2[validos] - v1[validos]
    n_red  = int((diff < 0).sum())
    n_nc   = int((v2[validos] == 0).sum())   # no consumió en IRT2
    n_ig   = int((diff == 0).sum())
    n_sub  = int((diff > 0).sum())
    evol_rows.append({
        'Sustancia': sust, 'N con sust. ppal': n_pp,
        'No consumió en IRT2': n_nc,
        'Redujo consumo': n_red,
        'Mismo consumo': n_ig,
        'Aumentó consumo': n_sub,
        'N válido': n_val,
    })
T23 = pd.DataFrame(evol_rows) if evol_rows else None

# ── 2.4 Cambio de consumo — formato TOP (hoja separada) ─────────────────────
# Solo consumidores en IRT1 (días > 0). Clasifica: Abstinencia / Disminuyó / Sin cambio / Empeoró
cambio_rows = []
for sust in SUST_ACTIVAS:
    entry = SUST_TOTAL[sust]
    c1 = entry.get('_IRT1'); c2 = entry.get('_IRT2')
    if not c1 or not c2: continue
    v1 = pd.to_numeric(df.loc[mask2, c1], errors='coerce')
    v2 = pd.to_numeric(df.loc[mask2, c2], errors='coerce')
    # Solo consumidores en IRT1 (>0)
    cons_irt1 = (v1 > 0) & v1.notna() & v2.notna()
    n_cons = int(cons_irt1.sum())
    if n_cons == 0: continue
    v1c = v1[cons_irt1]; v2c = v2[cons_irt1]
    n_abs  = int((v2c == 0).sum())
    n_dism = int(((v2c > 0) & (v2c < v1c)).sum())
    n_sc   = int((v2c == v1c).sum())
    n_emp  = int((v2c > v1c).sum())
    pct = lambda n: round(n / n_cons * 100, 1) if n_cons else 0.0
    cambio_rows.append({
        'Sustancia':       sust,
        'n cons. IRT1':    n_cons,
        'Abstinencia n':   n_abs,  'Abstinencia %':   pct(n_abs),
        'Disminuyó n':     n_dism, 'Disminuyó %':     pct(n_dism),
        'Sin cambio n':    n_sc,   'Sin cambio %':     pct(n_sc),
        'Empeoró n':       n_emp,  'Empeoró %':        pct(n_emp),
        '% Abs+Disminuyó': round((n_abs + n_dism) / n_cons * 100, 1) if n_cons else 0.0,
    })
T24 = pd.DataFrame(cambio_rows) if cambio_rows else None

# Fila TOTAL para T24
if T24 is not None and len(T24) > 0:
    tot_cons = T24['n cons. IRT1'].sum()
    tot_abs  = T24['Abstinencia n'].sum()
    tot_dism = T24['Disminuyó n'].sum()
    tot_sc   = T24['Sin cambio n'].sum()
    tot_emp  = T24['Empeoró n'].sum()
    pct_t = lambda n: round(n / tot_cons * 100, 1) if tot_cons else 0.0
    T24_total = {
        'Sustancia': 'TOTAL (todas las sustancias)',
        'n cons. IRT1': tot_cons,
        'Abstinencia n': tot_abs,   'Abstinencia %':   pct_t(tot_abs),
        'Disminuyó n':  tot_dism,   'Disminuyó %':     pct_t(tot_dism),
        'Sin cambio n': tot_sc,     'Sin cambio %':    pct_t(tot_sc),
        'Empeoró n':    tot_emp,    'Empeoró %':       pct_t(tot_emp),
        '% Abs+Disminuyó': round((tot_abs + tot_dism) / tot_cons * 100, 1) if tot_cons else 0.0,
    }
else:
    T24_total = None

# ── 2.5 % Consumidores por sustancia (IRT1 vs IRT2) ─────────────────────────
cons_rows = []
for sust in SUST_ACTIVAS:
    entry = SUST_TOTAL[sust]
    row = {'Sustancia': sust}
    for sfx, msk, N_sfx, lbl in [('_IRT1',mask2,N_irt2,'IRT1'),
                                   ('_IRT2',mask2,N_irt2,'IRT2'),
                                   ('_IRT3',mask3,N_irt3,'IRT3')]:
        if sfx == '_IRT3' and not TIENE_IRT3: continue
        c = entry.get(sfx)
        if c:
            vals = pd.to_numeric(df.loc[msk, c], errors='coerce').fillna(0)
            n_c  = int((vals > 0).sum())
            row[f'n {lbl}']   = n_c
            row[f'% {lbl}']   = round(n_c/N_sfx*100,1) if N_sfx else 0.0
        else:
            row[f'n {lbl}'] = '—'; row[f'% {lbl}'] = '—'
    cons_rows.append(row)
T25 = pd.DataFrame(cons_rows)

# ── 2.6 Promedio días consumo por sustancia (IRT1 vs IRT2) ──────────────────
prom_rows = []
for sust in SUST_ACTIVAS:
    entry = SUST_TOTAL[sust]
    row = {'Sustancia': sust}
    for sfx, msk, lbl in [('_IRT1',mask2,'IRT1'), ('_IRT2',mask2,'IRT2'),
                            ('_IRT3',mask3,'IRT3')]:
        if sfx == '_IRT3' and not TIENE_IRT3: continue
        c = entry.get(sfx)
        if c:
            vals = pd.to_numeric(df.loc[msk, c], errors='coerce')
            cons = vals[vals > 0].dropna()
            row[f'Prom. {lbl}'] = round(float(cons.mean()),1) if len(cons) else '—'
            row[f'N {lbl}']     = len(cons)
        else:
            row[f'Prom. {lbl}'] = '—'; row[f'N {lbl}'] = '—'
    prom_rows.append(row)
T26 = pd.DataFrame(prom_rows)

# ── 3.1 Urgencias ─────────────────────────────────────────────────────────────
urg_data = {}
for sfx, msk in [('_IRT1',mask2), ('_IRT2',mask2), ('_IRT3',mask3)]:
    if sfx == '_IRT3' and not TIENE_IRT3: continue
    nsi, pct, nv = calc_sino(COL_URG.get(sfx), msk)
    urg_data[sfx] = (nsi, pct, nv)

# ── 3.2 Accidentes ────────────────────────────────────────────────────────────
acc_data = {}
for sfx, msk in [('_IRT1',mask2), ('_IRT2',mask2), ('_IRT3',mask3)]:
    if sfx == '_IRT3' and not TIENE_IRT3: continue
    nsi, pct, nv = calc_sino(COL_ACC.get(sfx), msk)
    acc_data[sfx] = (nsi, pct, nv)

# ── 3.3 Salud ─────────────────────────────────────────────────────────────────
salud_rows = []
for nombre, cols_dict in [('Salud Psicológica (0–10)', COL_SPSI),
                           ('Salud Física (0–10)',       COL_SFIS)]:
    row = {'Dimensión': nombre}
    for sfx, msk, lbl in [('_IRT1',mask2,'IRT1'), ('_IRT2',mask2,'IRT2'),
                            ('_IRT3',mask3,'IRT3')]:
        if sfx == '_IRT3' and not TIENE_IRT3: continue
        m, nv = prom_col(cols_dict.get(sfx), msk)
        row[f'Prom. {lbl}'] = m if not np.isnan(m) else '—'
        row[f'N {lbl}']     = nv
    salud_rows.append(row)
T31 = pd.DataFrame(salud_rows)

# ── 3.4 Trabajo ───────────────────────────────────────────────────────────────
trab_rows = []
for nombre, cols_dict in [
    ('Tardanzas / salida anticipada', COL_TARDE),
    ('Faltas',                        COL_FALTAS),
    ('Sanciones / reprimendas',       COL_SANCIONES)]:
    row = {'Pregunta': nombre}
    for sfx, msk, lbl in [('_IRT1',mask2,'IRT1'), ('_IRT2',mask2,'IRT2'),
                            ('_IRT3',mask3,'IRT3')]:
        if sfx == '_IRT3' and not TIENE_IRT3: continue
        m, nv = prom_col(cols_dict.get(sfx), msk)
        row[f'Prom. {lbl}'] = m if not np.isnan(m) else '—'
        row[f'N {lbl}']     = nv
    trab_rows.append(row)
T34 = pd.DataFrame(trab_rows)

# ── 4. Transgresión ───────────────────────────────────────────────────────────
# Total con alguna transgresión
trans_total = {}
for sfx, msk in [('_IRT1',mask2), ('_IRT2',mask2), ('_IRT3',mask3)]:
    if sfx == '_IRT3' and not TIENE_IRT3: continue
    cols_tr = [d.get(sfx) for d in TRANS_COLS.values() if d.get(sfx)]
    if cols_tr:
        mask_any = pd.concat(
            [pd.to_numeric(df.loc[msk, c], errors='coerce').fillna(0) > 0
             for c in cols_tr], axis=1).any(axis=1)
        n_si = int(mask_any.sum())
        N_sfx = N_irt2 if sfx != '_IRT3' else N_irt3
        trans_total[sfx] = (n_si, round(n_si/N_sfx*100,1), N_sfx)

# Por tipo
trans_tipo_rows = []
for tipo, cols_dict in TRANS_COLS.items():
    row = {'Tipo': tipo}
    for sfx, msk, lbl in [('_IRT1',mask2,'IRT1'), ('_IRT2',mask2,'IRT2'),
                            ('_IRT3',mask3,'IRT3')]:
        if sfx == '_IRT3' and not TIENE_IRT3: continue
        nsi, pct, nv = n_trans(cols_dict.get(sfx), msk)
        row[f'n {lbl}'] = nsi; row[f'% {lbl}'] = pct
    trans_tipo_rows.append(row)
T42 = pd.DataFrame(trans_tipo_rows) if trans_tipo_rows else None

# ── 5. Relaciones interpersonales ─────────────────────────────────────────────
CATS_REL = ['Excelente', 'Buena', 'Ni buena ni mala', 'Mala', 'Muy mala']
rel_rows = []
for vinculo, cols_dict in REL_COLS.items():
    row = {'Vínculo': vinculo}
    for sfx, msk, lbl in [('_IRT1',mask2,'IRT1'), ('_IRT2',mask2,'IRT2'),
                            ('_IRT3',mask3,'IRT3')]:
        if sfx == '_IRT3' and not TIENE_IRT3: continue
        c = cols_dict.get(sfx)
        if c:
            vals = df.loc[msk, c].dropna()
            vals = vals[vals.astype(str).str.lower() != 'no aplica']
            nv   = len(vals)
            row[f'N aplica {lbl}'] = nv
            for cat in CATS_REL:
                nc = int((vals.astype(str).str.lower() == cat.lower()).sum())
                row[f'{cat} {lbl}'] = f'{nc} ({round(nc/nv*100,1) if nv else 0}%)'
    rel_rows.append(row)
T51 = pd.DataFrame(rel_rows) if rel_rows else None

# ── 6. Satisfacción ───────────────────────────────────────────────────────────
sat_rows = []
for dim, cols_dict in SAT_COLS.items():
    row = {'Dimensión': dim}
    for sfx, msk, lbl in [('_IRT1',mask2,'IRT1'), ('_IRT2',mask2,'IRT2'),
                            ('_IRT3',mask3,'IRT3')]:
        if sfx == '_IRT3' and not TIENE_IRT3: continue
        m, nv = prom_col(cols_dict.get(sfx), msk)
        row[f'Prom. {lbl}'] = m if not np.isnan(m) else '—'
        row[f'N {lbl}']     = nv
    sat_rows.append(row)
T61 = pd.DataFrame(sat_rows) if sat_rows else None

print('  ✓ Tablas calculadas')

# ══════════════════════════════════════════════════════════════════════════════
# ESCRITURA EXCEL
# ══════════════════════════════════════════════════════════════════════════════
print('\n→ Generando Excel...')
wb = Workbook()
ws = wb.active
ws.title = 'Seguimiento'
ws.sheet_properties.tabColor = C_IRT2
ws.sheet_view.showGridLines   = False

ws.column_dimensions['A'].width = 2
ws.column_dimensions['B'].width = 44
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 12
ws.column_dimensions['F'].width = 12
ws.column_dimensions['G'].width = 12
ws.column_dimensions['H'].width = 12

medium = Side(style='medium', color=C_MID)
thin   = Side(style='thin',   color='E0E8F0')

ROW = 1

# ── Helpers ────────────────────────────────────────────────────────────────────
def titulo_seccion(row, texto, subtexto=None):
    ws.row_dimensions[row].height = 26
    ws.merge_cells(f'B{row}:H{row}')
    c = ws.cell(row, 2, texto)
    c.font = Font(bold=True, size=12, color=C_WHITE, name='Arial')
    c.fill = PatternFill('solid', start_color=C_DARK)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    row += 1
    if subtexto:
        ws.row_dimensions[row].height = 14
        ws.merge_cells(f'B{row}:H{row}')
        c = ws.cell(row, 2, subtexto)
        c.font = Font(italic=True, size=8, color=C_NOTE, name='Arial')
        c.fill = PatternFill('solid', start_color='F2F6FC')
        c.alignment = Alignment(horizontal='left', vertical='center', indent=2)
        row += 1
    return row

def titulo_tabla(row, numero, texto):
    ws.row_dimensions[row].height = 20
    ws.merge_cells(f'B{row}:H{row}')
    c = ws.cell(row, 2, f'  {numero}  {texto}')
    c.font = Font(bold=True, size=10, color=C_WHITE, name='Arial')
    c.fill = PatternFill('solid', start_color=C_MID)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    return row + 1

def nota(row, texto):
    ws.row_dimensions[row].height = 14
    ws.merge_cells(f'B{row}:H{row}')
    c = ws.cell(row, 2, f'  {texto}')
    c.font = Font(italic=True, size=8, color=C_NOTE, name='Arial')
    c.fill = PatternFill('solid', start_color='F9FBFE')
    c.alignment = Alignment(horizontal='left', vertical='center', indent=2)
    return row + 1

def encabezados(row, hdrs):
    ws.row_dimensions[row].height = 18
    for ci, hdr in enumerate(hdrs, 2):
        c = ws.cell(row, ci, hdr)
        c.font = Font(bold=True, size=9, color=C_DARK, name='Arial')
        c.fill = PatternFill('solid', start_color=C_LIGHT)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = Border(bottom=medium)
    return row + 1

def fila_dato(row, label, valores, alt=False, negrita=False):
    ws.row_dimensions[row].height = 16
    bg = C_ALT if alt else C_WHITE
    c = ws.cell(row, 2, label)
    c.font = Font(size=9, name='Arial', bold=negrita)
    c.fill = PatternFill('solid', start_color=bg)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    c.border = Border(bottom=Side(style='thin', color='E0E8F0'))
    for ci, v in enumerate(valores, 3):
        c = ws.cell(row, ci, v)
        c.font = Font(size=9, name='Arial')
        c.fill = PatternFill('solid', start_color=bg)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = Border(bottom=Side(style='thin', color='E0E8F0'))
    return row + 1

def fila_total(row, label, valores):
    ws.row_dimensions[row].height = 18
    c = ws.cell(row, 2, label)
    c.font = Font(bold=True, size=9, color=C_WHITE, name='Arial')
    c.fill = PatternFill('solid', start_color=C_DARK)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    for ci, v in enumerate(valores, 3):
        c = ws.cell(row, ci, v)
        c.font = Font(bold=True, size=9, color=C_WHITE, name='Arial')
        c.fill = PatternFill('solid', start_color=C_DARK)
        c.alignment = Alignment(horizontal='center', vertical='center')
    return row + 1

def fila_cambio(row, label, valores, delta=None, alt=False):
    """Fila con coloreado semáforo en la columna Δ"""
    row = fila_dato(row-1, label, valores, alt=alt)
    # Colorear celda Δ si existe
    if delta is not None and not (isinstance(delta, float) and np.isnan(delta)):
        col_delta = 2 + len(valores)  # última columna de valores
        c = ws.cell(row-1, col_delta, delta)
        if isinstance(delta, (int, float)):
            if delta < 0:
                c.font = Font(size=9, name='Arial', bold=True, color=C_SUBE)
            elif delta > 0:
                c.font = Font(size=9, name='Arial', bold=True, color=C_BAJA)
            else:
                c.font = Font(size=9, name='Arial', color=C_IGUAL)
    return row

def spacer(row, n=1):
    for _ in range(n):
        ws.row_dimensions[row].height = 8
        row += 1
    return row

# ── Etiquetas para encabezados según IRT3 ────────────────────────────────────
LBL = ['IRT1\n(Ingreso)', 'IRT2\n(Seg. 3m)']
if TIENE_IRT3: LBL.append('IRT3\n(Seg. 6m)')

# ══════════════════════════════════════════════════════════════════════════════
# ENCABEZADO PRINCIPAL
# ══════════════════════════════════════════════════════════════════════════════
ws.row_dimensions[ROW].height = 36
ws.merge_cells(f'B{ROW}:H{ROW}')
c = ws.cell(ROW, 2,
    f'INFORME DE SEGUIMIENTO  ·  IRT  ·  {SERVICIO}  ·  {PERIODO}')
c.font = Font(bold=True, size=14, color=C_WHITE, name='Arial')
c.fill = PatternFill('solid', start_color=C_IRT2)
c.alignment = Alignment(horizontal='left', vertical='center', indent=2)
ROW += 1

ws.row_dimensions[ROW].height = 18
ws.merge_cells(f'B{ROW}:H{ROW}')
c = ws.cell(ROW, 2,
    f'N total = {N_total}  ·  Con IRT2: {N_irt2}  ·  Con IRT3: {N_irt3}'
    f'  ·  Análisis sobre pacientes con IRT2 (n={N_irt2})')
c.font = Font(italic=True, size=9, color=C_NOTE, name='Arial')
c.fill = PatternFill('solid', start_color='EEF4FB')
c.alignment = Alignment(horizontal='left', vertical='center', indent=2)
ROW = spacer(ROW + 1, 2)

# ══════════════════════════════════════════════════════════════════════════════
# SECCIÓN 0 — INSTRUMENTOS CONTESTADOS
# ══════════════════════════════════════════════════════════════════════════════
ROW = titulo_seccion(ROW, '0.  INSTRUMENTOS CONTESTADOS')
ROW = titulo_tabla(ROW, '0.1', 'Número de Participantes por Etapa de Aplicación')
ROW = encabezados(ROW, ['', 'Instrumento', 'n', '% del total'])
for lbl_inst, n_inst in [
    ('IRT 1 — Ingreso a tratamiento',          N_total),
    ('IRT 2 — Seguimiento a los 3 meses',      N_irt2),
    ('IRT 3 — Seguimiento a los 6 meses',      N_irt3),
]:
    ROW = fila_dato(ROW, '', [lbl_inst, n_inst,
                               f'{round(n_inst/N_total*100,1)}%'])
ROW = spacer(ROW, 2)

# ══════════════════════════════════════════════════════════════════════════════
# SECCIÓN 1 — SOCIODEMOGRAFÍA
# ══════════════════════════════════════════════════════════════════════════════
ROW = titulo_seccion(ROW,
    '1.  ANTECEDENTES GENERALES',
    f'Análisis sobre los {N_irt2} participantes con seguimiento (IRT2)')

if T11 is not None:
    ROW = titulo_tabla(ROW, '1.1', 'Distribución por Sexo')
    ROW = encabezados(ROW, ['Sexo', 'n', '%'])
    for i, r in T11.iterrows():
        ROW = fila_dato(ROW, r['Categoría'], [r['n'], f"{r['%']}%"], alt=i%2==0)
    ROW = fila_total(ROW, 'Total', [N_irt2, '100%'])
    ROW = spacer(ROW, 2)

if T12 is not None:
    ROW = titulo_tabla(ROW, '1.2', 'Distribución por Rango de Edad')
    if edad_m:
        ROW = nota(ROW, f'Edad promedio: {edad_m:.1f} años  (mín: {edad_min:.0f} – máx: {edad_max:.0f})')
    ROW = encabezados(ROW, ['Rango', 'n', '%'])
    for i, r in T12.iterrows():
        ROW = fila_dato(ROW, r['Rango'], [r['n'], f"{r['%']}%"], alt=i%2==0)
    ROW = fila_total(ROW, 'Total', [T12['n'].sum(), '100%'])
    ROW = spacer(ROW, 2)

# ══════════════════════════════════════════════════════════════════════════════
# SECCIÓN 2 — CONSUMO DE SUSTANCIAS
# ══════════════════════════════════════════════════════════════════════════════
ROW = titulo_seccion(ROW,
    '2.  CONSUMO DE SUSTANCIAS',
    'Comparación IRT1 (ingreso) vs IRT2 (seguimiento 3 meses)'
    + (' vs IRT3 (seguimiento 6 meses)' if TIENE_IRT3 else ''))

# 2.1 Sustancia principal
ROW = titulo_tabla(ROW, '2.1', 'Sustancia Principal (IRT1 vs IRT2)')
ROW = nota(ROW, 'Sustancia declarada como la que genera más problemas al momento de cada aplicación.')
hdrs_21 = ['Sustancia', 'n IRT1', '% IRT1', 'n IRT2', '% IRT2']
if TIENE_IRT3: hdrs_21 += ['n IRT3', '% IRT3']
ROW = encabezados(ROW, hdrs_21)
for i, r in T21.iterrows():
    vals = [r['n IRT1'], f"{r['% IRT1']}%", r['n IRT2'], f"{r['% IRT2']}%"]
    if TIENE_IRT3: vals += [r['n IRT3'], f"{r['% IRT3']}%"]
    ROW = fila_dato(ROW, r['Sustancia'], vals, alt=i%2==0)
ROW = spacer(ROW, 2)

# 2.2 Promedio días sustancia principal
if T22 is not None and len(T22) > 0:
    ROW = titulo_tabla(ROW, '2.2', 'Promedio de Días de Consumo — Sustancia Principal')
    ROW = nota(ROW, 'Calculado sobre las personas cuya sustancia principal es cada una de las indicadas.')
    ROW = encabezados(ROW, ['Sustancia', 'Etapa', 'Prom. días (0–28)', 'n'])
    for i, r in T22.iterrows():
        ROW = fila_dato(ROW, r['Sustancia'], [r['Etapa'], r['Promedio días (0–28)'], r['n']], alt=i%2==0)
    ROW = spacer(ROW, 2)

# 2.3 Evolución consumo sustancia principal
if T23 is not None and len(T23) > 0:
    ROW = titulo_tabla(ROW, '2.3', 'Evolución del Consumo — Sustancia Principal (IRT1 → IRT2)')
    ROW = nota(ROW, 'Analiza si las personas redujeron, mantuvieron o aumentaron el consumo de su sustancia principal entre IRT1 e IRT2.')
    ROW = encabezados(ROW, ['Sustancia', 'N con sust. ppal', 'No consumió IRT2', 'Redujo', 'Mismo consumo', 'Aumentó', 'N válido'])
    for i, r in T23.iterrows():
        pct_nc  = round(r['No consumió en IRT2']/r['N válido']*100,1) if r['N válido'] else 0
        pct_red = round(r['Redujo consumo']/r['N válido']*100,1)       if r['N válido'] else 0
        pct_sub = round(r['Aumentó consumo']/r['N válido']*100,1)      if r['N válido'] else 0
        ROW = fila_dato(ROW, r['Sustancia'],
                        [r['N con sust. ppal'],
                         f"{r['No consumió en IRT2']} ({pct_nc}%)",
                         f"{r['Redujo consumo']} ({pct_red}%)",
                         f"{r['Mismo consumo']}",
                         f"{r['Aumentó consumo']} ({pct_sub}%)",
                         r['N válido']], alt=i%2==0)
    ROW = spacer(ROW, 2)

# 2.4 Cambio de consumo → referencia a hoja separada
if T24 is not None and len(T24) > 0:
    ROW = titulo_tabla(ROW, '2.4', 'Cambio de Consumo por Sustancia (IRT1 → IRT2)')
    ROW = nota(ROW, '→ Ver hoja "Cambio de Consumo" para el detalle completo (Abstinencia / Disminuyó / Sin cambio / Empeoró).')
    ROW = spacer(ROW, 2)

# 2.5 % Consumidores
ROW = titulo_tabla(ROW, '2.5', 'Porcentaje de Consumidores por Sustancia (IRT1 vs IRT2)')
ROW = nota(ROW, 'Un participante puede consumir más de una sustancia. Puede superar 100%.')
hdrs_25 = ['Sustancia', 'n IRT1', '% IRT1', 'n IRT2', '% IRT2']
if TIENE_IRT3: hdrs_25 += ['n IRT3', '% IRT3']
ROW = encabezados(ROW, hdrs_25)
for i, r in T25.iterrows():
    vals = [r['n IRT1'], f"{r['% IRT1']}%" if r['% IRT1'] != '—' else '—',
            r['n IRT2'], f"{r['% IRT2']}%" if r['% IRT2'] != '—' else '—']
    if TIENE_IRT3:
        vals += [r['n IRT3'], f"{r['% IRT3']}%" if r['% IRT3'] != '—' else '—']
    ROW = fila_dato(ROW, r['Sustancia'], vals, alt=i%2==0)
ROW = spacer(ROW, 2)

# 2.6 Promedio días por sustancia
ROW = titulo_tabla(ROW, '2.6', 'Promedio de Días de Consumo por Sustancia (IRT1 vs IRT2)')
ROW = nota(ROW, 'Calculado solo sobre consumidores (días > 0) en cada etapa.')
hdrs_26 = ['Sustancia', 'Prom. IRT1', 'N IRT1', 'Prom. IRT2', 'N IRT2']
if TIENE_IRT3: hdrs_26 += ['Prom. IRT3', 'N IRT3']
ROW = encabezados(ROW, hdrs_26)
for i, r in T26.iterrows():
    vals = [r['Prom. IRT1'], r['N IRT1'], r['Prom. IRT2'], r['N IRT2']]
    if TIENE_IRT3: vals += [r['Prom. IRT3'], r['N IRT3']]
    ROW = fila_dato(ROW, r['Sustancia'], vals, alt=i%2==0)
ROW = spacer(ROW, 2)

# ══════════════════════════════════════════════════════════════════════════════
# SECCIÓN 3 — SALUD Y FUNCIONAMIENTO
# ══════════════════════════════════════════════════════════════════════════════
ROW = titulo_seccion(ROW,
    '3.  SALUD Y FUNCIONAMIENTO',
    'Comparación de indicadores de salud entre IRT1 e IRT2')

# 3.1 Urgencias
if any(v[0] is not None for v in urg_data.values()):
    ROW = titulo_tabla(ROW, '3.1', 'Urgencia u Hospitalización por Consumo (IRT1 vs IRT2)')
    ROW = nota(ROW, '¿Ha acudido a urgencias u hospitalización relacionada con el consumo en las últimas 4 semanas?')
    hdrs_31 = ['', 'n IRT1', '% IRT1', 'n IRT2', '% IRT2']
    if TIENE_IRT3: hdrs_31 += ['n IRT3', '% IRT3']
    ROW = encabezados(ROW, hdrs_31)
    for lbl_r, key in [('Con urgencia/hospitalización', True), ('Sin urgencia/hospitalización', False)]:
        vals = []
        for sfx in (['_IRT1','_IRT2'] + (['_IRT3'] if TIENE_IRT3 else [])):
            d = urg_data.get(sfx, (None,None,None))
            if d[0] is None:
                vals += ['—','—']; continue
            nsi, pct, nv = d
            v = nsi if key else (nv - nsi)
            p = pct if key else round((nv-nsi)/nv*100,1) if nv else 0
            vals += [v, f'{p}%']
        ROW = fila_dato(ROW, lbl_r, vals, alt=not key)
    ROW = spacer(ROW, 1)

# 3.2 Accidentes
if any(v[0] is not None for v in acc_data.values()):
    ROW = titulo_tabla(ROW, '3.2', 'Accidentes por Consumo (IRT1 vs IRT2)')
    ROW = nota(ROW, '¿Ha tenido algún accidente relacionado con su consumo en las últimas 4 semanas?')
    hdrs_32 = ['', 'n IRT1', '% IRT1', 'n IRT2', '% IRT2']
    if TIENE_IRT3: hdrs_32 += ['n IRT3', '% IRT3']
    ROW = encabezados(ROW, hdrs_32)
    for lbl_r, key in [('Con accidente por consumo', True), ('Sin accidente por consumo', False)]:
        vals = []
        for sfx in (['_IRT1','_IRT2'] + (['_IRT3'] if TIENE_IRT3 else [])):
            d = acc_data.get(sfx, (None,None,None))
            if d[0] is None:
                vals += ['—','—']; continue
            nsi, pct, nv = d
            v = nsi if key else (nv - nsi)
            p = pct if key else round((nv-nsi)/nv*100,1) if nv else 0
            vals += [v, f'{p}%']
        ROW = fila_dato(ROW, lbl_r, vals, alt=not key)
    ROW = spacer(ROW, 1)

# 3.3 Salud
ROW = titulo_tabla(ROW, '3.3', 'Autopercepción del Estado de Salud (IRT1 vs IRT2)')
ROW = nota(ROW, 'Escala de 0 a 10 — donde 0 = muy mala y 10 = excelente.')
hdrs_33 = ['Dimensión', 'Prom. IRT1', 'N IRT1', 'Prom. IRT2', 'N IRT2']
if TIENE_IRT3: hdrs_33 += ['Prom. IRT3', 'N IRT3']
ROW = encabezados(ROW, hdrs_33)
for i, r in T31.iterrows():
    vals = [r['Prom. IRT1'], r['N IRT1'], r['Prom. IRT2'], r['N IRT2']]
    if TIENE_IRT3: vals += [r['Prom. IRT3'], r['N IRT3']]
    ROW = fila_dato(ROW, r['Dimensión'], vals, alt=i%2==0)
ROW = spacer(ROW, 1)

# 3.4 Trabajo
if len(T34) > 0:
    ROW = titulo_tabla(ROW, '3.4', 'Problemas en Trabajo o Institución Educacional (IRT1 vs IRT2)')
    ROW = nota(ROW, 'Promedio de veces que ocurrió cada situación en las últimas 4 semanas.')
    hdrs_34 = ['Pregunta', 'Prom. IRT1', 'N IRT1', 'Prom. IRT2', 'N IRT2']
    if TIENE_IRT3: hdrs_34 += ['Prom. IRT3', 'N IRT3']
    ROW = encabezados(ROW, hdrs_34)
    for i, r in T34.iterrows():
        vals = [r['Prom. IRT1'], r['N IRT1'], r['Prom. IRT2'], r['N IRT2']]
        if TIENE_IRT3: vals += [r['Prom. IRT3'], r['N IRT3']]
        ROW = fila_dato(ROW, r['Pregunta'], vals, alt=i%2==0)
    ROW = spacer(ROW, 2)

# ══════════════════════════════════════════════════════════════════════════════
# SECCIÓN 4 — TRANSGRESIÓN
# ══════════════════════════════════════════════════════════════════════════════
if T42 is not None:
    ROW = titulo_seccion(ROW,
        '4.  TRANSGRESIÓN A LA NORMA SOCIAL',
        'Comparación entre IRT1 e IRT2')

    ROW = titulo_tabla(ROW, '4.1', 'Personas con al Menos una Conducta de Transgresión (IRT1 vs IRT2)')
    hdrs_41 = ['', 'n IRT1', '% IRT1', 'n IRT2', '% IRT2']
    if TIENE_IRT3: hdrs_41 += ['n IRT3', '% IRT3']
    ROW = encabezados(ROW, hdrs_41)
    for lbl_r, key in [('Con transgresión', True), ('Sin transgresión', False)]:
        vals = []
        for sfx in (['_IRT1','_IRT2'] + (['_IRT3'] if TIENE_IRT3 else [])):
            d = trans_total.get(sfx)
            if d is None:
                vals += ['—','—']; continue
            nsi, pct, N_sfx = d
            v = nsi if key else (N_sfx - nsi)
            p = pct if key else round((N_sfx-nsi)/N_sfx*100,1) if N_sfx else 0
            vals += [v, f'{p}%']
        ROW = fila_dato(ROW, lbl_r, vals, alt=not key)
    ROW = spacer(ROW, 1)

    ROW = titulo_tabla(ROW, '4.2', 'Distribución por Tipo de Transgresión (IRT1 vs IRT2)')
    ROW = nota(ROW, 'Porcentaje calculado sobre el total de participantes con IRT2. Puede superar 100%.')
    hdrs_42 = ['Tipo de Transgresión', 'n IRT1', '% IRT1', 'n IRT2', '% IRT2']
    if TIENE_IRT3: hdrs_42 += ['n IRT3', '% IRT3']
    ROW = encabezados(ROW, hdrs_42)
    for i, r in T42.iterrows():
        vals = [r['n IRT1'], f"{r['% IRT1']}%", r['n IRT2'], f"{r['% IRT2']}%"]
        if TIENE_IRT3: vals += [r['n IRT3'], f"{r['% IRT3']}%"]
        ROW = fila_dato(ROW, r['Tipo'], vals, alt=i%2==0)
    ROW = spacer(ROW, 2)

# ══════════════════════════════════════════════════════════════════════════════
# SECCIÓN 5 — RELACIONES INTERPERSONALES
# ══════════════════════════════════════════════════════════════════════════════
if T51 is not None and len(T51) > 0:
    CATS_PRES = [cat for cat in CATS_REL if any(
        f'{cat} IRT1' in T51.columns or f'{cat} IRT2' in T51.columns for _ in [None])]
    ROW = titulo_seccion(ROW,
        '5.  RELACIONES INTERPERSONALES',
        'Calidad de relaciones en IRT1 vs IRT2. Excluye respuestas "No aplica".')
    ROW = titulo_tabla(ROW, '5.1', 'Distribución por Vínculo (IRT1 vs IRT2)')
    # Encabezados con IRT1 e IRT2 separados
    hdrs_51 = ['Vínculo', 'N IRT1']
    for cat in CATS_REL: hdrs_51.append(f'{cat} IRT1')
    hdrs_51.append('N IRT2')
    for cat in CATS_REL: hdrs_51.append(f'{cat} IRT2')
    # Versión simplificada: solo N aplica + Excelente+Buena vs Mala+Muy mala
    ROW = nota(ROW, 'Se muestra: N válido y distribución por categoría en cada etapa.')
    ROW = encabezados(ROW, ['Vínculo', 'N IRT1', 'Excelente\nIRT1', 'Buena\nIRT1',
                              'N IRT2', 'Excelente\nIRT2', 'Buena\nIRT2'])
    for i, r in T51.iterrows():
        n1 = r.get('N aplica IRT1', '—')
        n2 = r.get('N aplica IRT2', '—')
        e1 = r.get('Excelente IRT1', '—'); b1 = r.get('Buena IRT1', '—')
        e2 = r.get('Excelente IRT2', '—'); b2 = r.get('Buena IRT2', '—')
        ROW = fila_dato(ROW, r['Vínculo'], [n1, e1, b1, n2, e2, b2], alt=i%2==0)
    ROW = spacer(ROW, 2)

# ══════════════════════════════════════════════════════════════════════════════
# SECCIÓN 6 — SATISFACCIÓN DE VIDA
# ══════════════════════════════════════════════════════════════════════════════
if T61 is not None and len(T61) > 0:
    ROW = titulo_seccion(ROW,
        '6.  SATISFACCIÓN DE VIDA',
        'Escala de 0 a 10 — donde 0 = nada satisfecho y 10 = muy satisfecho')
    ROW = titulo_tabla(ROW, '6.1', 'Promedio de Satisfacción por Dimensión (IRT1 vs IRT2)')
    hdrs_61 = ['Dimensión', 'Prom. IRT1', 'N IRT1', 'Prom. IRT2', 'N IRT2']
    if TIENE_IRT3: hdrs_61 += ['Prom. IRT3', 'N IRT3']
    ROW = encabezados(ROW, hdrs_61)
    for i, r in T61.iterrows():
        vals = [r['Prom. IRT1'], r['N IRT1'], r['Prom. IRT2'], r['N IRT2']]
        if TIENE_IRT3: vals += [r['Prom. IRT3'], r['N IRT3']]
        ROW = fila_dato(ROW, r['Dimensión'], vals, alt=i%2==0)
    ROW = spacer(ROW, 2)

# ── Nota al pie ───────────────────────────────────────────────────────────────
ws.row_dimensions[ROW].height = 14
ws.merge_cells(f'B{ROW}:H{ROW}')
c = ws.cell(ROW, 2,
    f'  Fuente: {os.path.basename(INPUT_FILE)}  ·  '
    f'Generado automáticamente · SCRIPT_IRT_Universal_Seguimiento_Excel v1.1')
c.font = Font(italic=True, size=7, color='AAAAAA', name='Arial')
c.alignment = Alignment(horizontal='left', vertical='center', indent=1)

# ══════════════════════════════════════════════════════════════════════════════
# HOJA SEPARADA — CAMBIO DE CONSUMO
# ══════════════════════════════════════════════════════════════════════════════
if T24 is not None and len(T24) > 0:
    wc = wb.create_sheet('Cambio de Consumo')
    wc.sheet_properties.tabColor = C_IRT2
    wc.sheet_view.showGridLines   = False

    # Anchos
    wc.column_dimensions['A'].width = 2
    wc.column_dimensions['B'].width = 30
    wc.column_dimensions['C'].width = 11
    wc.column_dimensions['D'].width = 11
    wc.column_dimensions['E'].width = 11
    wc.column_dimensions['F'].width = 11
    wc.column_dimensions['G'].width = 11
    wc.column_dimensions['H'].width = 11
    wc.column_dimensions['I'].width = 11
    wc.column_dimensions['J'].width = 11
    wc.column_dimensions['K'].width = 13

    RC = 1  # puntero fila

    # Título principal
    wc.row_dimensions[RC].height = 28
    wc.merge_cells(f'B{RC}:K{RC}')
    c = wc.cell(RC, 2,
        f'CAMBIO EN EL CONSUMO POR SUSTANCIA  ·  IRT  ·  Ingreso → Seguimiento')
    c.font      = Font(bold=True, size=13, color=C_WHITE, name='Arial')
    c.fill      = PatternFill('solid', start_color=C_IRT2)
    c.alignment = Alignment(horizontal='center', vertical='center')
    RC += 1

    # Subtítulo
    wc.row_dimensions[RC].height = 16
    wc.merge_cells(f'B{RC}:K{RC}')
    c = wc.cell(RC, 2,
        f'  Solo pacientes con consumo > 0 en IRT1  ·  '
        f'% sobre n consumidores  ·  N seguimiento = {N_irt2}')
    c.font      = Font(italic=True, size=9, color=C_NOTE, name='Arial')
    c.fill      = PatternFill('solid', start_color='EEF4FB')
    c.alignment = Alignment(horizontal='left', vertical='center', indent=2)
    RC += 1

    # Espacio
    wc.row_dimensions[RC].height = 6; RC += 1

    # Encabezados dobles — fila 1
    COLS_CC = ['B','C','D','E','F','G','H','I','J','K']
    hdrs1 = ['Sustancia', 'n cons.\nIRT1',
             'Abstinencia\nn', 'Abstinencia\n%',
             'Disminuyó\nn',   'Disminuyó\n%',
             'Sin cambio\nn',  'Sin cambio\n%',
             'Empeoró\nn',     'Empeoró\n%',
             '% Abs +\nDisminuyó']
    wc.row_dimensions[RC].height = 30
    for col_l, hdr in zip(COLS_CC, hdrs1):
        c = wc.cell(RC, ord(col_l)-64, hdr)
        c.font      = Font(bold=True, size=9, color=C_DARK, name='Arial')
        c.fill      = PatternFill('solid', start_color=C_LIGHT)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border    = Border(bottom=Side(style='medium', color=C_MID))
    RC += 1

    # Filas de datos
    C_ABS  = '2E75B6'   # azul — abstinencia
    C_DISM = '375623'   # verde — disminuyó
    C_SC   = '595959'   # gris — sin cambio
    C_EMP  = 'C00000'   # rojo — empeoró
    C_PORC = '1F3864'   # azul oscuro — % Abs+Dism

    for i, r in T24.iterrows():
        alt = i % 2 == 0
        bg  = C_ALT if alt else C_WHITE
        wc.row_dimensions[RC].height = 17

        vals_cols = [
            (r['Sustancia'],         C_DARK, False),
            (r['n cons. IRT1'],      C_DARK, True),
            (r['Abstinencia n'],     C_ABS,  True),
            (f"{r['Abstinencia %']}%",  C_ABS, True),
            (r['Disminuyó n'],       C_DISM, True),
            (f"{r['Disminuyó %']}%",    C_DISM, True),
            (r['Sin cambio n'],      C_SC,   True),
            (f"{r['Sin cambio %']}%",   C_SC, True),
            (r['Empeoró n'],         C_EMP,  True),
            (f"{r['Empeoró %']}%",      C_EMP, True),
            (f"{r['% Abs+Disminuyó']}%", C_PORC, True),
        ]
        for ci, (val, color, centrado) in enumerate(vals_cols, 2):
            c = wc.cell(RC, ci, val)
            c.font      = Font(size=9, name='Arial', color=color)
            c.fill      = PatternFill('solid', start_color=bg)
            c.alignment = Alignment(
                horizontal='center' if centrado else 'left',
                vertical='center',
                indent=0 if centrado else 1)
            c.border = Border(bottom=Side(style='thin', color='E0E8F0'))
        RC += 1

    # Fila TOTAL
    if T24_total:
        wc.row_dimensions[RC].height = 18
        tot_vals = [
            (T24_total['Sustancia'],             C_WHITE, False),
            (T24_total['n cons. IRT1'],          C_WHITE, True),
            (T24_total['Abstinencia n'],         C_WHITE, True),
            (f"{T24_total['Abstinencia %']}%",   C_WHITE, True),
            (T24_total['Disminuyó n'],           C_WHITE, True),
            (f"{T24_total['Disminuyó %']}%",     C_WHITE, True),
            (T24_total['Sin cambio n'],          C_WHITE, True),
            (f"{T24_total['Sin cambio %']}%",    C_WHITE, True),
            (T24_total['Empeoró n'],             C_WHITE, True),
            (f"{T24_total['Empeoró %']}%",       C_WHITE, True),
            (f"{T24_total['% Abs+Disminuyó']}%", C_WHITE, True),
        ]
        for ci, (val, color, centrado) in enumerate(tot_vals, 2):
            c = wc.cell(RC, ci, val)
            c.font      = Font(bold=True, size=9, name='Arial', color=color)
            c.fill      = PatternFill('solid', start_color=C_DARK)
            c.alignment = Alignment(
                horizontal='center' if centrado else 'left',
                vertical='center',
                indent=0 if centrado else 1)
        RC += 1

    # Nota metodológica
    RC += 1
    wc.row_dimensions[RC].height = 14
    wc.merge_cells(f'B{RC}:K{RC}')
    c = wc.cell(RC, 2,
        '  Abstinencia = consumo 0 en IRT2 (con consumo > 0 en IRT1). '
        'Disminuyó = IRT2 < IRT1. Sin cambio = IRT2 = IRT1. Empeoró = IRT2 > IRT1. '
        '% sobre n consumidores de esa sustancia en IRT1.')
    c.font      = Font(italic=True, size=8, color=C_NOTE, name='Arial')
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)

    print('  ✓ Hoja "Cambio de Consumo" generada')

wb.save(OUTPUT_FILE)
print(f'\n{"=" * 60}')
print(f'  ✅  LISTO  →  {OUTPUT_FILE}')
print(f'      {N_irt2} pacientes con IRT2  ·  {PERIODO}')
print(f'{"=" * 60}')
