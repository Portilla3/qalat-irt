"""
╔══════════════════════════════════════════════════════════════════════════════╗
║   SCRIPT_IRT_Universal_Caracterizacion_Excel.py  —  v1.1                  ║
║   Genera tablas de caracterización al ingreso (IRT1)                       ║
║   Compatible con cualquier país que use el instrumento IRT                 ║
╠══════════════════════════════════════════════════════════════════════════════╣
║  CÓMO USAR:                                                                 ║
║  1. Abre un chat nuevo con Claude                                           ║
║  2. Sube DOS archivos:                                                      ║
║       • Este script                                                         ║
║       • La base Wide IRT (generada por SCRIPT_IRT_Universal_Wide)          ║
║  3. Escribe: "Ejecuta el script Caracterización IRT con esta base Wide"    ║
║                                                                             ║
║  TABLAS GENERADAS (solo IRT1 — ingreso):                                   ║
║    1.1  Distribución por Sexo                                               ║
║    1.2  Distribución por Rango de Edad                                      ║
║    2.1  Distribución Sustancia Principal                                    ║
║    2.2  Promedio Días de Consumo por Sustancia Principal                    ║
║    2.3  Consumo de Sustancias (% consumidores, puede >100%)                ║
║    2.4  Promedio Días de Consumo por Sustancia                             ║
║    3.1  Autopercepción Salud Psicológica y Física (0–10)                   ║
║    4.1  Transgresión a la Norma Social                                      ║
║    4.2  Tipos de Transgresión                                               ║
║    5.1  Relaciones Interpersonales (por vínculo)                           ║
║    6.1  Satisfacción de Vida (por dimensión)                               ║
╚══════════════════════════════════════════════════════════════════════════════╝
"""
import glob, os, unicodedata, warnings
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
warnings.filterwarnings('ignore')

# ── Colores ───────────────────────────────────────────────────────────────────
C_DARK  = '1F3864'; C_MID   = '2E75B6'; C_LIGHT = 'BDD7EE'
C_ALT   = 'EEF4FB'; C_WHITE = 'FFFFFF'; C_BDR   = 'B8CCE4'
C_NOTE  = '595959'; C_IRT2  = '00B0F0'

def _norm(s):
    return unicodedata.normalize('NFD', str(s).lower()).encode('ascii','ignore').decode()

# ── Auto-detección de archivo Wide ───────────────────────────────────────────
def auto_archivo_wide():
    candidatos = (
        glob.glob('/mnt/user-data/uploads/*IRT*Wide*.xlsx') +
        glob.glob('/mnt/user-data/uploads/*Wide*IRT*.xlsx') +
        glob.glob('/mnt/user-data/uploads/IRT_Base*.xlsx') +
        glob.glob('/mnt/user-data/outputs/IRT_Base_Wide*.xlsx') +
        glob.glob('/home/claude/IRT_Base_Wide.xlsx'))
    if not candidatos:
        raise FileNotFoundError(
            "\n⚠  No se encontró la base Wide IRT.\n"
            "   Sube el archivo IRT_Base_Wide.xlsx junto con este script.")
    # Priorizar uploads; si no, el de mayor tamaño (más pacientes)
    uploads = [f for f in candidatos if 'uploads' in f]
    if uploads:
        elegido = uploads[0]
    else:
        elegido = max(candidatos, key=os.path.getsize)
    print(f"  → Base Wide detectada: {os.path.basename(elegido)}")
    return elegido

# ══════════════════════════════════════════════════════════════════════════════
print('=' * 60)
print('  SCRIPT_IRT_Universal_Caracterizacion_Excel  v1.1')
print('=' * 60)

INPUT_FILE  = auto_archivo_wide()

# ── FILTRO POR CENTRO (opcional) ──────────────────────────────────────────────
#   FILTRO_CENTRO = None        ← todos los centros
#   FILTRO_CENTRO = "FUNCADREDEF2"  ← solo ese centro
FILTRO_CENTRO = None
# ─────────────────────────────────────────────────────────────────────────────

OUTPUT_FILE = '/home/claude/IRT_Caracterizacion_Ingreso.xlsx'

df = pd.read_excel(INPUT_FILE, sheet_name='Base Wide', header=1)
df.columns = [str(c) for c in df.columns]
cols = df.columns.tolist()

N_total = len(df)
print(f'\n→ {N_total} pacientes cargados')
print(f'  IRT2: {(df["Tiene_IRT2"]=="Sí").sum()} | IRT3: {(df["Tiene_IRT3"]=="Sí").sum()}')

# ── Detectar período y servicio ───────────────────────────────────────────────
MESES_ES = {1:'Enero',2:'Febrero',3:'Marzo',4:'Abril',5:'Mayo',6:'Junio',
            7:'Julio',8:'Agosto',9:'Septiembre',10:'Octubre',
            11:'Noviembre',12:'Diciembre'}

fecha_col = next((c for c in cols if 'fecha de administracion' in _norm(c)
                  or 'fecha_aplicacion' in _norm(c)), None)
if fecha_col:
    fechas = pd.to_datetime(df[fecha_col], errors='coerce').dropna()
    anio_actual = pd.Timestamp.now().year
    fechas = fechas[(fechas.dt.year >= anio_actual - 10) &
                    (fechas.dt.year <= anio_actual + 1)]
    if len(fechas):
        f0, f1 = fechas.min(), fechas.max()
        if f0.year == f1.year and f0.month == f1.month:
            PERIODO = f'{MESES_ES[f0.month]} {f0.year}'
        elif f0.year == f1.year:
            PERIODO = f'{MESES_ES[f0.month]}–{MESES_ES[f1.month]} {f0.year}'
        else:
            PERIODO = f'{MESES_ES[f0.month]} {f0.year} – {MESES_ES[f1.month]} {f1.year}'
    else:
        PERIODO = 'Período no determinado'
else:
    PERIODO = 'Período no determinado'

centro_col = next((c for c in cols if any(x in _norm(c) for x in
                   ['codigo del centro', 'servicio de tratamiento',
                    'centro/ servicio', 'codigo centro'])), None)

# Aplicar filtro por centro si está definido
if FILTRO_CENTRO and centro_col:
    n_antes = len(df)
    df = df[df[centro_col].astype(str).str.strip() == FILTRO_CENTRO].copy()
    df = df.reset_index(drop=True)
    cols = df.columns.tolist()
    print(f'  ⚑ Filtro activo: Centro = "{FILTRO_CENTRO}"')
    print(f'    {n_antes} pacientes totales → {len(df)} del centro seleccionado')
    OUTPUT_FILE = f'/home/claude/IRT_Caracterizacion_{FILTRO_CENTRO}.xlsx'
elif FILTRO_CENTRO and not centro_col:
    print(f'  ⚠ FILTRO_CENTRO = "{FILTRO_CENTRO}" pero no se encontró columna de centro.')

if centro_col:
    centros = df[centro_col].dropna().unique()
    SERVICIO = centros[0] if len(centros) == 1 else f'{len(centros)} centros'
else:
    SERVICIO = 'Servicio de Tratamiento'

print(f'  Período : {PERIODO}')
print(f'  Servicio: {SERVICIO}')

# ══════════════════════════════════════════════════════════════════════════════
# DETECTAR COLUMNAS CLAVE (sufijo _IRT1)
# ══════════════════════════════════════════════════════════════════════════════
def col_irt1(keywords, obligatoria=True):
    """Busca columna _IRT1 que contenga todas las keywords"""
    for c in cols:
        if not c.endswith('_IRT1'): continue
        nc = _norm(c)
        if all(_norm(k) in nc for k in keywords):
            return c
    if obligatoria:
        raise ValueError(f'Columna IRT1 no encontrada: {keywords}')
    return None

def cols_irt1_multi(keyword):
    """Retorna todas las columnas _IRT1 que contienen keyword"""
    return [c for c in cols if c.endswith('_IRT1') and _norm(keyword) in _norm(c)]

# Identificación
COL_FN     = next((c for c in cols if 'fecha de nacimiento' in _norm(c)
                   or 'fecha_nac' in _norm(c)), None)
COL_SEXO   = next((c for c in cols if _norm(c) in ['sexo','género','genero']), None)

# Consumo — detectar sustancias disponibles
SUST_NOMBRES = {
    'Alcohol':        ['alcohol'],
    'Marihuana':      ['marihuana','cannabis','marijuana'],
    'Heroína':        ['heroina','heroin'],
    'Cocaína':        ['cocain','cocaína'],
    'Fentanilo':      ['fentanil'],
    'Inhalables':     ['inhalab','inhalant'],
    'Metanfetamina':  ['metanfet','cristal','anfetam'],
    'Crack':          ['crack'],
    'Pasta Base':     ['pasta base','pasta'],
    'Sedantes':       ['sedant','tranquil','benzod'],
    'Opiáceos':       ['opiod','opiac','morfin'],
    'Tabaco':         ['tabaco','cigarr','nicot'],
    'Otra sustancia': ['otra sust','otrasust'],
}

# Mapear sustancias a sus columnas _total_IRT1 y _cantidad_IRT1
SUST_COLS_TOTAL    = {}  # sust → col total días (0-28)
SUST_COLS_CANTIDAD = {}  # sust → col cantidad/promedio

for sust, kws in SUST_NOMBRES.items():
    for c in cols:
        if not c.endswith('_IRT1'): continue
        nc = _norm(c)
        if not any(_norm(k) in nc for k in kws): continue
        if 'total' in nc or '(0-28)' in nc:
            SUST_COLS_TOTAL[sust]    = c
        elif 'cantidad' in nc or 'promedio' in nc:
            SUST_COLS_CANTIDAD[sust] = c

# Solo sustancias con datos reales
SUST_ACTIVAS = [s for s in SUST_NOMBRES if s in SUST_COLS_TOTAL]
print(f'\n→ Sustancias detectadas: {SUST_ACTIVAS}')

# Sustancia principal
COL_SUST_PPAL = col_irt1(['sustancia','principal'], obligatoria=False)

# Salud
COL_SALUD_PSI = col_irt1(['salud','psicol'], obligatoria=False)
COL_SALUD_FIS = col_irt1(['salud','fis'],    obligatoria=False)

# Urgencias / Hospitalización
COL_URGENCIA  = next((c for c in cols if c.endswith('_IRT1') and
                      '5)' in c and any(k in c.lower() for k in
                      ['urgencia','hospitali','emergencia'])), None)
COL_URG_CANT  = next((c for c in cols if c.endswith('_IRT1') and
                      '5.1)' in c and any(k in c.lower() for k in
                      ['urgencia','veces','hospitali'])), None)

# Accidentes
COL_ACCIDENTE = next((c for c in cols if c.endswith('_IRT1') and
                      '6)' in c and 'accidente' in c.lower()), None)
COL_ACC_CANT  = next((c for c in cols if c.endswith('_IRT1') and
                      '6.1)' in c and any(k in c.lower() for k in
                      ['accidente','veces'])), None)

# Trabajo / Educación
COL_TARDE     = next((c for c in cols if c.endswith('_IRT1') and
                      '10)' in c and any(k in c.lower() for k in
                      ['tarde','atrasado','antes de su jornada'])), None)
COL_FALTAS    = next((c for c in cols if c.endswith('_IRT1') and
                      '10)' in c and 'falt' in c.lower()), None)
COL_SANCIONES = next((c for c in cols if c.endswith('_IRT1') and
                      '10)' in c and 'sancion' in c.lower()), None)

print(f'  Urgencias : {COL_URGENCIA is not None} | Accidentes: {COL_ACCIDENTE is not None}')
print(f'  Trabajo   : tarde={COL_TARDE is not None} faltas={COL_FALTAS is not None} sanciones={COL_SANCIONES is not None}')

# Transgresión
COL_ROBO      = col_irt1(['robo'],              obligatoria=False)
COL_VENTA     = col_irt1(['venta'],             obligatoria=False)
COL_VIF       = col_irt1(['violencia','intraf'], obligatoria=False)
COL_VIO_OTRAS = col_irt1(['violencia','otras'],  obligatoria=False)
COL_DETENIDO  = col_irt1(['detenido'],           obligatoria=False)

# Relaciones
REL_COLS = {}
REL_MAP  = {
    'Padre':   ['padre'],    'Madre':  ['madre'],
    'Hijos':   ['hijos','hijo'], 'Hermanos': ['hermanos','hermano'],
    'Pareja':  ['pareja'],   'Amigos': ['amigos','amigo'],
    'Otros':   ['otros'],
}
for vinculo, kws in REL_MAP.items():
    for c in cols:
        if not c.endswith('_IRT1'): continue
        nc = _norm(c)
        if '14)' not in c and 'relaci' not in nc: continue
        if any(_norm(k) in nc for k in kws):
            REL_COLS[vinculo] = c; break

print(f'  Vínculos relacionales: {list(REL_COLS.keys())}')

# Satisfacción
SAT_MAP = {
    'Vida en general':     [['16)'], ['satisfac','vida','disfrutar']],
    'Lugar donde vive':    [['17)'], ['satisfac','lugar','vive']],
    'Situación laboral':   [['18)'], ['satisfac','labor','educac']],
    'Tiempo libre':        [['19)'], ['satisfac','tiempo libre','actividad']],
    'Capacidad económica': [['20)'], ['satisfac','econom','responsab']],
}
SAT_COLS = {}
for dim, (nums, kws) in SAT_MAP.items():
    for c in cols:
        if not c.endswith('_IRT1'): continue
        nc = _norm(c)
        if any(num in c for num in nums) and any(_norm(k) in nc for k in kws):
            SAT_COLS[dim] = c; break
    if dim not in SAT_COLS:
        # fallback: solo por número de pregunta
        for c in cols:
            if not c.endswith('_IRT1'): continue
            if any(f'{num} ' in c or c.startswith(num) for num in nums):
                SAT_COLS[dim] = c; break

print(f'  Dimensiones satisfacción: {list(SAT_COLS.keys())}')

# ══════════════════════════════════════════════════════════════════════════════
# CÁLCULOS
# ══════════════════════════════════════════════════════════════════════════════
print('\n→ Calculando tablas...')

# ── 1.1 Sexo ──────────────────────────────────────────────────────────────────
if COL_SEXO:
    sexo = df[COL_SEXO].value_counts(dropna=True)
    T11 = pd.DataFrame({'Categoría': sexo.index,
                        'n': sexo.values,
                        '%': (sexo.values / N_total * 100).round(1)})
else:
    T11 = pd.DataFrame({'Categoría': ['Sin dato'], 'n': [N_total], '%': [100.0]})

# ── 1.2 Edad ─────────────────────────────────────────────────────────────────
hoy = pd.Timestamp.now()
if COL_FN:
    df['_edad'] = ((hoy - pd.to_datetime(df[COL_FN], errors='coerce')).dt.days / 365.25)
    bins   = [0, 17, 25, 35, 45, 55, 200]
    labels = ['< 18 años', '18–25 años', '26–35 años',
              '36–45 años', '46–55 años', '56 años o más']
    df['_grupo_edad'] = pd.cut(df['_edad'], bins=bins, labels=labels)
    edad_dist = df['_grupo_edad'].value_counts().reindex(labels, fill_value=0)
    T12 = pd.DataFrame({'Rango': labels,
                        'n': edad_dist.values,
                        '%': (edad_dist.values / N_total * 100).round(1)})
    edad_mean = df['_edad'].mean()
    edad_min  = df['_edad'].min()
    edad_max  = df['_edad'].max()
else:
    T12 = None
    edad_mean = edad_min = edad_max = None

# ── 2.1 Sustancia principal ───────────────────────────────────────────────────
def norm_sust(s):
    if pd.isna(s) or str(s).strip() in ['0','']: return None
    s = str(s).strip().lower()
    s = unicodedata.normalize('NFD', s).encode('ascii','ignore').decode()
    if any(x in s for x in ['alcohol','cerveza','licor','aguard','bebida']): return 'Alcohol'
    if any(x in s for x in ['marihu','cannabis','marij','hierba','mota']):   return 'Marihuana'
    if any(x in s for x in ['crack','piedra','paco']):                        return 'Crack'
    if any(x in s for x in ['pasta base','pasta','papelillo']):               return 'Pasta Base'
    if any(x in s for x in ['cocain','perico','coca']):                       return 'Cocaína'
    if any(x in s for x in ['fentanil']):                                     return 'Fentanilo'
    if any(x in s for x in ['inhalab','thiner','activo','cemento']):          return 'Inhalables'
    if any(x in s for x in ['metanfet','cristal','anfetam']):                 return 'Metanfetamina'
    if any(x in s for x in ['sedant','benzod','tranquil','valium','clonaz']): return 'Sedantes'
    if any(x in s for x in ['opiod','heroina','morfin','fentanil']):          return 'Opiáceos'
    if any(x in s for x in ['tabaco','cigarr','nicot']):                      return 'Tabaco'
    return 'Otra sustancia'

if COL_SUST_PPAL:
    sp = df[COL_SUST_PPAL].apply(norm_sust).dropna()
    sp_cnt = sp.value_counts()
    T21 = pd.DataFrame({'Sustancia': sp_cnt.index,
                        'n': sp_cnt.values,
                        '%': (sp_cnt.values / len(sp) * 100).round(1)})
else:
    T21 = None

# ── 2.1b Promedio días sustancia principal ───────────────────────────────────
# Para cada persona: tomar su sustancia principal y buscar sus días totales
ppal_prom_rows = []
if COL_SUST_PPAL:
    df['_sust_norm'] = df[COL_SUST_PPAL].apply(norm_sust)
    for sust in SUST_ACTIVAS:
        mask    = df['_sust_norm'] == sust
        n_ppal  = int(mask.sum())
        if n_ppal == 0: continue
        c_tot   = SUST_COLS_TOTAL.get(sust)
        if c_tot is None: continue
        dias    = pd.to_numeric(df.loc[mask, c_tot], errors='coerce').dropna()
        if len(dias) == 0: continue
        ppal_prom_rows.append({
            'Sustancia principal': sust,
            'n': n_ppal,
            'Promedio días (0–28)': round(dias.mean(), 1),
        })
    # Ordenar por n
    T21b = pd.DataFrame(ppal_prom_rows).sort_values('n', ascending=False) if ppal_prom_rows else None
else:
    T21b = None

# ── 2.2 Consumidores por sustancia ────────────────────────────────────────────
consumo_rows = []
for sust in SUST_ACTIVAS:
    c_tot = SUST_COLS_TOTAL.get(sust)
    if not c_tot: continue
    vals = pd.to_numeric(df[c_tot], errors='coerce')
    n_cons = int((vals > 0).sum())
    pct    = round(n_cons / N_total * 100, 1)
    consumo_rows.append({'Sustancia': sust, 'n': n_cons, '%': pct})
T22 = pd.DataFrame(consumo_rows).sort_values('%', ascending=False)

# ── 2.3 Promedio días consumo ─────────────────────────────────────────────────
prom_rows = []
for sust in SUST_ACTIVAS:
    c_tot = SUST_COLS_TOTAL.get(sust)
    if not c_tot: continue
    vals = pd.to_numeric(df[c_tot], errors='coerce')
    cons = vals[vals > 0]
    if len(cons) == 0: continue
    prom_rows.append({
        'Sustancia': sust,
        'Promedio días (solo consumidores)': round(cons.mean(), 1),
        'N válido': len(cons),
    })
T23 = pd.DataFrame(prom_rows).sort_values('Promedio días (solo consumidores)', ascending=False)

# ── 3.0 Urgencias / Hospitalización ──────────────────────────────────────────
def calc_sino(col):
    """Para columnas Sí/No retorna (n_si, pct_si, n_valido)"""
    if col is None: return None, None, None
    vals = df[col].dropna().astype(str).str.strip().str.lower()
    n_val = len(vals)
    n_si  = int(vals.isin(['sí','si','yes','1','true']).sum())
    return n_si, round(n_si / n_val * 100, 1) if n_val else 0.0, n_val

n_urg, pct_urg, nv_urg   = calc_sino(COL_URGENCIA)
n_acc, pct_acc, nv_acc   = calc_sino(COL_ACCIDENTE)

# Promedio veces (solo quienes dijeron Sí)
def prom_si(col_sino, col_cant):
    if col_sino is None or col_cant is None: return None, None
    mask = df[col_sino].astype(str).str.strip().str.lower().isin(['sí','si','yes','1','true'])
    vals = pd.to_numeric(df.loc[mask, col_cant], errors='coerce').dropna()
    return round(vals.mean(), 1) if len(vals) else 0.0, len(vals)

prom_urg, n_prom_urg = prom_si(COL_URGENCIA, COL_URG_CANT)
prom_acc, n_prom_acc = prom_si(COL_ACCIDENTE, COL_ACC_CANT)

# ── 3.1 Salud ─────────────────────────────────────────────────────────────────
salud_rows = []
for nombre, col in [('Salud Psicológica (0–10)', COL_SALUD_PSI),
                    ('Salud Física (0–10)',       COL_SALUD_FIS)]:
    if col:
        vals = pd.to_numeric(df[col], errors='coerce').dropna()
        salud_rows.append({
            'Dimensión': nombre,
            'Promedio': round(vals.mean(), 1),
            'Mínimo': int(vals.min()),
            'Máximo': int(vals.max()),
            'N válido': len(vals),
        })
T31 = pd.DataFrame(salud_rows) if salud_rows else None

# ── 3.x Trabajo / Educación ──────────────────────────────────────────────────
trabajo_rows = []
for nombre, col in [
    ('¿Cuántas veces llegó tarde o se fue antes de su jornada?', COL_TARDE),
    ('¿Cuántas veces faltó a su trabajo o institución educativa?', COL_FALTAS),
    ('¿Cuántas veces fue sancionado/reprimido/advertido?',        COL_SANCIONES),
]:
    if col is None: continue
    vals = pd.to_numeric(df[col], errors='coerce').dropna()
    trabajo_rows.append({
        'Pregunta':  nombre,
        'Promedio':  round(vals.mean(), 1) if len(vals) else '—',
        'N válido':  len(vals),
    })
T_trabajo = pd.DataFrame(trabajo_rows) if trabajo_rows else None

# ── 4.1 Transgresión ─────────────────────────────────────────────────────────
# Columnas binarias o numéricas → al menos 1 vez = Sí
def cuenta_transgresores(col):
    if col is None: return 0, 0.0
    vals = pd.to_numeric(df[col], errors='coerce').fillna(0)
    n = int((vals > 0).sum())
    return n, round(n / N_total * 100, 1)

n_trans_alg = 0
transgresion_rows = []
for nombre, col in [
    ('Robo / Hurto',             COL_ROBO),
    ('Venta de sustancias',      COL_VENTA),
    ('Violencia a otras personas', COL_VIO_OTRAS),
    ('Violencia intrafamiliar',  COL_VIF),
    ('Detenido / Arrestado',     COL_DETENIDO),
]:
    if col is None: continue
    n, pct = cuenta_transgresores(col)
    transgresion_rows.append({'Tipo': nombre, 'n': n, '%': pct})

# Total con al menos un tipo
if transgresion_rows:
    cols_trans = [c for c in [COL_ROBO, COL_VENTA, COL_VIO_OTRAS,
                               COL_VIF, COL_DETENIDO] if c]
    mask_trans = pd.concat(
        [pd.to_numeric(df[c], errors='coerce').fillna(0) > 0 for c in cols_trans],
        axis=1).any(axis=1)
    n_trans_alg = int(mask_trans.sum())

T41 = pd.DataFrame(transgresion_rows) if transgresion_rows else None

# ── 5.1 Relaciones interpersonales ───────────────────────────────────────────
CATS_REL  = ['Excelente', 'Buena', 'Ni buena ni mala', 'Mala', 'Muy mala']
rel_rows  = []
for vinculo, col in REL_COLS.items():
    validos = df[col].dropna()
    validos = validos[validos.astype(str).str.lower() != 'no aplica']
    n_val   = len(validos)
    if n_val == 0: continue
    row = {'Vínculo': vinculo, 'N aplica': n_val}
    for cat in CATS_REL:
        n_cat = int((validos.astype(str).str.lower() == cat.lower()).sum())
        row[cat] = f'{n_cat} ({round(n_cat/n_val*100,1)}%)'
    rel_rows.append(row)
T51 = pd.DataFrame(rel_rows) if rel_rows else None

# ── 6.1 Satisfacción de vida ─────────────────────────────────────────────────
sat_rows = []
for dim, col in SAT_COLS.items():
    vals = pd.to_numeric(df[col], errors='coerce').dropna()
    if len(vals) == 0: continue
    sat_rows.append({
        'Dimensión': dim,
        'Promedio (0–10)': round(vals.mean(), 1),
        'Mínimo': int(vals.min()),
        'Máximo': int(vals.max()),
        'N válido': len(vals),
    })
T61 = pd.DataFrame(sat_rows) if sat_rows else None

print('  ✓ Tablas calculadas')

# ══════════════════════════════════════════════════════════════════════════════
# ESCRITURA EXCEL
# ══════════════════════════════════════════════════════════════════════════════
print('\n→ Generando Excel...')
wb  = Workbook()
ws  = wb.active
ws.title = 'Caracterización Ingreso'
ws.sheet_properties.tabColor = C_DARK
ws.sheet_view.showGridLines   = False

# Anchos de columna
ws.column_dimensions['A'].width = 2
ws.column_dimensions['B'].width = 40
ws.column_dimensions['C'].width = 14
ws.column_dimensions['D'].width = 14
ws.column_dimensions['E'].width = 14
ws.column_dimensions['F'].width = 14
ws.column_dimensions['G'].width = 14
ws.column_dimensions['H'].width = 14

thin   = Side(style='thin',   color=C_BDR)
medium = Side(style='medium', color=C_MID)

ROW = 1  # puntero de fila

# ── Helpers de formato ────────────────────────────────────────────────────────
def titulo_seccion(row, texto, subtexto=None):
    ws.row_dimensions[row].height = 26
    ws.merge_cells(f'B{row}:H{row}')
    c = ws.cell(row, 2, texto)
    c.font      = Font(bold=True, size=12, color=C_WHITE, name='Arial')
    c.fill      = PatternFill('solid', start_color=C_DARK)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    row += 1
    if subtexto:
        ws.row_dimensions[row].height = 14
        ws.merge_cells(f'B{row}:H{row}')
        c = ws.cell(row, 2, subtexto)
        c.font      = Font(italic=True, size=8, color=C_NOTE, name='Arial')
        c.fill      = PatternFill('solid', start_color='F2F6FC')
        c.alignment = Alignment(horizontal='left', vertical='center', indent=2)
        row += 1
    return row

def titulo_tabla(row, numero, texto):
    ws.row_dimensions[row].height = 20
    ws.merge_cells(f'B{row}:H{row}')
    c = ws.cell(row, 2, f'  {numero}  {texto}')
    c.font      = Font(bold=True, size=10, color=C_WHITE, name='Arial')
    c.fill      = PatternFill('solid', start_color=C_MID)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    return row + 1

def encabezados(row, hdrs, anchos_col='CDEFGH'):
    ws.row_dimensions[row].height = 18
    for ci, (col_l, hdr) in enumerate(zip(anchos_col, hdrs)):
        c = ws.cell(row, ord(col_l) - 64, hdr)
        c.font      = Font(bold=True, size=9, color=C_DARK, name='Arial')
        c.fill      = PatternFill('solid', start_color=C_LIGHT)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border    = Border(bottom=medium)
    return row + 1

def fila_dato(row, label, valores, alt=False, negrita_label=False, indent=1):
    ws.row_dimensions[row].height = 16
    bg = C_ALT if alt else C_WHITE
    c = ws.cell(row, 2, label)
    c.font      = Font(size=9, name='Arial', bold=negrita_label)
    c.fill      = PatternFill('solid', start_color=bg)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=indent)
    c.border    = Border(bottom=Side(style='thin', color='E0E8F0'))
    for ci, v in enumerate(valores, 3):
        c = ws.cell(row, ci, v)
        c.font      = Font(size=9, name='Arial')
        c.fill      = PatternFill('solid', start_color=bg)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border    = Border(bottom=Side(style='thin', color='E0E8F0'))
    return row + 1

def fila_total(row, label, valores):
    ws.row_dimensions[row].height = 18
    c = ws.cell(row, 2, label)
    c.font      = Font(bold=True, size=9, color=C_WHITE, name='Arial')
    c.fill      = PatternFill('solid', start_color=C_DARK)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    for ci, v in enumerate(valores, 3):
        c = ws.cell(row, ci, v)
        c.font      = Font(bold=True, size=9, color=C_WHITE, name='Arial')
        c.fill      = PatternFill('solid', start_color=C_DARK)
        c.alignment = Alignment(horizontal='center', vertical='center')
    return row + 1

def spacer(row, n=1):
    for _ in range(n):
        ws.row_dimensions[row].height = 8
        row += 1
    return row

# ══════════════════════════════════════════════════════════════════════════════
# ENCABEZADO PRINCIPAL
# ══════════════════════════════════════════════════════════════════════════════
ws.row_dimensions[ROW].height = 36
ws.merge_cells(f'B{ROW}:H{ROW}')
c = ws.cell(ROW, 2,
    f'CARACTERIZACIÓN AL INGRESO  ·  IRT  ·  {SERVICIO}  ·  {PERIODO}')
c.font      = Font(bold=True, size=14, color=C_WHITE, name='Arial')
c.fill      = PatternFill('solid', start_color=C_DARK)
c.alignment = Alignment(horizontal='left', vertical='center', indent=2)
ROW += 1

ws.row_dimensions[ROW].height = 18
ws.merge_cells(f'B{ROW}:H{ROW}')
c = ws.cell(ROW, 2,
    f'N = {N_total} pacientes al ingreso  ·  '
    f'IRT2: {(df["Tiene_IRT2"]=="Sí").sum()}  ·  '
    f'Fuente: {os.path.basename(INPUT_FILE)}')
c.font      = Font(italic=True, size=9, color=C_NOTE, name='Arial')
c.fill      = PatternFill('solid', start_color='EEF4FB')
c.alignment = Alignment(horizontal='left', vertical='center', indent=2)
ROW = spacer(ROW + 1, 2)

# ══════════════════════════════════════════════════════════════════════════════
# SECCIÓN 1 — CARACTERIZACIÓN SOCIODEMOGRÁFICA
# ══════════════════════════════════════════════════════════════════════════════
ROW = titulo_seccion(ROW,
    '1.  CARACTERIZACIÓN SOCIODEMOGRÁFICA',
    'Distribución de los participantes según sexo y edad al momento de ingreso a tratamiento')

# 1.1 Sexo
ROW = titulo_tabla(ROW, '1.1', 'Distribución por Sexo')
ROW = encabezados(ROW, ['Sexo', 'n', '%'])
for i, row_d in T11.iterrows():
    ROW = fila_dato(ROW, row_d['Categoría'], [row_d['n'], f"{row_d['%']}%"], alt=i%2==0)
ROW = fila_total(ROW, f'Total', [N_total, '100%'])
ROW = spacer(ROW, 2)

# 1.2 Edad
if T12 is not None:
    ROW = titulo_tabla(ROW, '1.2', 'Distribución por Rango de Edad')
    if edad_mean:
        nota_edad = (f'Edad promedio: {edad_mean:.1f} años  '
                     f'(mín: {edad_min:.0f} – máx: {edad_max:.0f})')
        ws.row_dimensions[ROW].height = 14
        ws.merge_cells(f'B{ROW}:H{ROW}')
        c = ws.cell(ROW, 2, f'  {nota_edad}')
        c.font      = Font(italic=True, size=8, color=C_NOTE, name='Arial')
        c.fill      = PatternFill('solid', start_color='F9FBFE')
        c.alignment = Alignment(horizontal='left', vertical='center', indent=2)
        ROW += 1
    ROW = encabezados(ROW, ['Rango de Edad', 'n', '%'])
    for i, row_d in T12.iterrows():
        ROW = fila_dato(ROW, row_d['Rango'], [row_d['n'], f"{row_d['%']}%"], alt=i%2==0)
    n_edad_valido = T12['n'].sum()
    ROW = fila_total(ROW, 'Total', [n_edad_valido, '100%'])
    ROW = spacer(ROW, 2)

# ══════════════════════════════════════════════════════════════════════════════
# SECCIÓN 2 — CONSUMO DE SUSTANCIAS
# ══════════════════════════════════════════════════════════════════════════════
ROW = titulo_seccion(ROW,
    '2.  CONSUMO DE SUSTANCIAS',
    'Basado en el reporte de consumo durante las últimas 4 semanas (28 días)')

# 2.1 Distribución sustancia principal
if T21 is not None:
    ROW = titulo_tabla(ROW, '2.1', 'Distribución según Sustancia Principal')
    ws.row_dimensions[ROW].height = 14
    ws.merge_cells(f'B{ROW}:H{ROW}')
    c = ws.cell(ROW, 2,
        '  Sustancia que el participante declara que le genera más problemas al ingresar a tratamiento.')
    c.font      = Font(italic=True, size=8, color=C_NOTE, name='Arial')
    c.fill      = PatternFill('solid', start_color='F9FBFE')
    c.alignment = Alignment(horizontal='left', vertical='center', indent=2)
    ROW += 1
    ROW = encabezados(ROW, ['Sustancia', 'n', '%'])
    for i, row_d in T21.iterrows():
        ROW = fila_dato(ROW, row_d['Sustancia'], [row_d['n'], f"{row_d['%']}%"], alt=i%2==0)
    ROW = fila_total(ROW, 'Total con dato', [T21['n'].sum(), '—'])
    ROW = spacer(ROW, 2)

# 2.2 Promedio días sustancia principal
if T21b is not None and len(T21b) > 0:
    ROW = titulo_tabla(ROW, '2.2',
        'Promedio de Días de Consumo por Sustancia Principal')
    ws.row_dimensions[ROW].height = 14
    ws.merge_cells(f'B{ROW}:H{ROW}')
    c = ws.cell(ROW, 2,
        '  Promedio de días consumidos en las últimas 4 semanas, '
        'agrupado según la sustancia principal declarada por cada persona. (n = personas con esa sustancia principal)')
    c.font      = Font(italic=True, size=8, color=C_NOTE, name='Arial')
    c.fill      = PatternFill('solid', start_color='F9FBFE')
    c.alignment = Alignment(horizontal='left', vertical='center', indent=2)
    ROW += 1
    ROW = encabezados(ROW, ['Sustancia Principal', 'n', 'Prom. días (0–28)'])
    for i, row_d in T21b.iterrows():
        ROW = fila_dato(ROW, row_d['Sustancia principal'],
                        [row_d['n'], row_d['Promedio días (0–28)']], alt=i%2==0)
    ROW = spacer(ROW, 2)

# 2.3 Consumidores por sustancia
ROW = titulo_tabla(ROW, '2.3',
    'Consumo de Sustancias (porcentaje de consumidores, puede superar 100%)')
ws.row_dimensions[ROW].height = 14
ws.merge_cells(f'B{ROW}:H{ROW}')
c = ws.cell(ROW, 2,
    f'  Un participante puede consumir más de una sustancia. '
    f'Porcentaje calculado sobre N = {N_total}.')
c.font      = Font(italic=True, size=8, color=C_NOTE, name='Arial')
c.fill      = PatternFill('solid', start_color='F9FBFE')
c.alignment = Alignment(horizontal='left', vertical='center', indent=2)
ROW += 1
ROW = encabezados(ROW, ['Sustancia', 'n consumidores', '% del total'])
for i, row_d in T22.iterrows():
    ROW = fila_dato(ROW, row_d['Sustancia'],
                    [row_d['n'], f"{row_d['%']}%"], alt=i%2==0)
ROW = spacer(ROW, 2)

# 2.4 Promedio días de consumo por sustancia (todas)
ROW = titulo_tabla(ROW, '2.4',
    'Promedio de Días de Consumo por Sustancia (solo consumidores)')
ws.row_dimensions[ROW].height = 14
ws.merge_cells(f'B{ROW}:H{ROW}')
c = ws.cell(ROW, 2,
    '  Promedio calculado únicamente sobre personas con consumo > 0 días en las últimas 4 semanas.')
c.font      = Font(italic=True, size=8, color=C_NOTE, name='Arial')
c.fill      = PatternFill('solid', start_color='F9FBFE')
c.alignment = Alignment(horizontal='left', vertical='center', indent=2)
ROW += 1
ROW = encabezados(ROW, ['Sustancia', 'Prom. días (0–28)', 'N válido'])
for i, row_d in T23.iterrows():
    ROW = fila_dato(ROW, row_d['Sustancia'],
                    [row_d['Promedio días (solo consumidores)'],
                     row_d['N válido']], alt=i%2==0)
ROW = spacer(ROW, 2)

# ══════════════════════════════════════════════════════════════════════════════
# SECCIÓN 3 — SALUD Y FUNCIONAMIENTO
# ══════════════════════════════════════════════════════════════════════════════
ROW = titulo_seccion(ROW,
    '3.  SALUD Y FUNCIONAMIENTO',
    'Indicadores de salud y situaciones ocurridas en las 4 semanas previas al ingreso')

# 3.1 Urgencias / Hospitalización
if n_urg is not None:
    ROW = titulo_tabla(ROW, '3.1', 'Urgencia u Hospitalización por Consumo de Sustancias')
    ws.row_dimensions[ROW].height = 14
    ws.merge_cells(f'B{ROW}:H{ROW}')
    c = ws.cell(ROW, 2,
        '  ¿Ha acudido a urgencias u hospitalización por consumo en las últimas 4 semanas?')
    c.font      = Font(italic=True, size=8, color=C_NOTE, name='Arial')
    c.fill      = PatternFill('solid', start_color='F9FBFE')
    c.alignment = Alignment(horizontal='left', vertical='center', indent=2)
    ROW += 1
    ROW = encabezados(ROW, ['', 'n', '%', 'Prom. veces (solo Sí)', 'N válido'])
    ROW = fila_dato(ROW, 'Con urgencia u hospitalización',
                    [n_urg, f'{pct_urg}%',
                     prom_urg if prom_urg is not None else '—', nv_urg])
    ROW = fila_dato(ROW, 'Sin urgencia u hospitalización',
                    [nv_urg - n_urg,
                     f'{round((nv_urg-n_urg)/nv_urg*100,1) if nv_urg else 0}%', '—', '—'],
                    alt=True)
    ROW = fila_total(ROW, 'Total', [nv_urg, '100%', '—', '—'])
    ROW = spacer(ROW, 1)

# 3.2 Accidentes
if n_acc is not None:
    ROW = titulo_tabla(ROW, '3.2', 'Accidentes Relacionados con el Consumo de Sustancias')
    ws.row_dimensions[ROW].height = 14
    ws.merge_cells(f'B{ROW}:H{ROW}')
    c = ws.cell(ROW, 2,
        '  ¿Ha tenido algún accidente relacionado con su consumo en las últimas 4 semanas?')
    c.font      = Font(italic=True, size=8, color=C_NOTE, name='Arial')
    c.fill      = PatternFill('solid', start_color='F9FBFE')
    c.alignment = Alignment(horizontal='left', vertical='center', indent=2)
    ROW += 1
    ROW = encabezados(ROW, ['', 'n', '%', 'Prom. veces (solo Sí)', 'N válido'])
    ROW = fila_dato(ROW, 'Con accidente por consumo',
                    [n_acc, f'{pct_acc}%',
                     prom_acc if prom_acc is not None else '—', nv_acc])
    ROW = fila_dato(ROW, 'Sin accidente por consumo',
                    [nv_acc - n_acc,
                     f'{round((nv_acc-n_acc)/nv_acc*100,1) if nv_acc else 0}%', '—', '—'],
                    alt=True)
    ROW = fila_total(ROW, 'Total', [nv_acc, '100%', '—', '—'])
    ROW = spacer(ROW, 1)

# 3.3 Autopercepción salud
if T31 is not None:
    ROW = titulo_tabla(ROW, '3.3', 'Autopercepción del Estado de Salud Psicológica y Física')
    ws.row_dimensions[ROW].height = 14
    ws.merge_cells(f'B{ROW}:H{ROW}')
    c = ws.cell(ROW, 2,
        '  Escala de 0 a 10 — donde 0 = muy mala y 10 = excelente.')
    c.font      = Font(italic=True, size=8, color=C_NOTE, name='Arial')
    c.fill      = PatternFill('solid', start_color='F9FBFE')
    c.alignment = Alignment(horizontal='left', vertical='center', indent=2)
    ROW += 1
    ROW = encabezados(ROW, ['Dimensión', 'Promedio', 'Mín.', 'Máx.', 'N válido'])
    for i, row_d in T31.iterrows():
        ROW = fila_dato(ROW, row_d['Dimensión'],
                        [row_d['Promedio'], row_d['Mínimo'],
                         row_d['Máximo'], row_d['N válido']], alt=i%2==0)
    ROW = spacer(ROW, 1)

# 3.4 Problemas en el trabajo o institución educacional
if T_trabajo is not None and len(T_trabajo) > 0:
    ROW = titulo_tabla(ROW, '3.4',
        'Problemas en el Trabajo o Institución Educacional')
    ws.row_dimensions[ROW].height = 14
    ws.merge_cells(f'B{ROW}:H{ROW}')
    c = ws.cell(ROW, 2,
        '  Número de veces que ocurrió cada situación en las últimas 4 semanas. '
        'El número entre paréntesis indica el N válido.')
    c.font      = Font(italic=True, size=8, color=C_NOTE, name='Arial')
    c.fill      = PatternFill('solid', start_color='F9FBFE')
    c.alignment = Alignment(horizontal='left', vertical='center', indent=2)
    ROW += 1
    ROW = encabezados(ROW, ['Pregunta', 'Promedio', 'N válido'])
    for i, row_d in T_trabajo.iterrows():
        ROW = fila_dato(ROW, row_d['Pregunta'],
                        [row_d['Promedio'], f"({row_d['N válido']})"],
                        alt=i%2==0)
    ROW = spacer(ROW, 2)

# ══════════════════════════════════════════════════════════════════════════════
# SECCIÓN 4 — TRANSGRESIÓN
# ══════════════════════════════════════════════════════════════════════════════
if T41 is not None:
    ROW = titulo_seccion(ROW,
        '4.  TRANSGRESIÓN A LA NORMA SOCIAL',
        'Número de veces que ocurrió cada conducta en las últimas 4 semanas o 2 meses')

    # 4.1 Total con transgresión
    ROW = titulo_tabla(ROW, '4.1', 'Personas con al Menos una Conducta de Transgresión')
    ROW = encabezados(ROW, ['', 'n', '%'])
    ROW = fila_dato(ROW, 'Con algún tipo de transgresión',
                    [n_trans_alg, f'{round(n_trans_alg/N_total*100,1)}%'])
    ROW = fila_dato(ROW, 'Sin transgresión',
                    [N_total - n_trans_alg,
                     f'{round((N_total-n_trans_alg)/N_total*100,1)}%'], alt=True)
    ROW = fila_total(ROW, 'Total', [N_total, '100%'])
    ROW = spacer(ROW, 1)

    # 4.2 Por tipo
    ROW = titulo_tabla(ROW, '4.2', 'Distribución por Tipo de Transgresión')
    ws.row_dimensions[ROW].height = 14
    ws.merge_cells(f'B{ROW}:H{ROW}')
    c = ws.cell(ROW, 2,
        '  Porcentaje calculado sobre el total de participantes. '
        'Una persona puede tener más de un tipo.')
    c.font      = Font(italic=True, size=8, color=C_NOTE, name='Arial')
    c.fill      = PatternFill('solid', start_color='F9FBFE')
    c.alignment = Alignment(horizontal='left', vertical='center', indent=2)
    ROW += 1
    ROW = encabezados(ROW, ['Tipo de Transgresión', 'n', '%'])
    for i, row_d in T41.iterrows():
        ROW = fila_dato(ROW, row_d['Tipo'],
                        [row_d['n'], f"{row_d['%']}%"], alt=i%2==0)
    ROW = spacer(ROW, 2)

# ══════════════════════════════════════════════════════════════════════════════
# SECCIÓN 5 — RELACIONES INTERPERSONALES
# ══════════════════════════════════════════════════════════════════════════════
if T51 is not None and len(T51) > 0:
    ROW = titulo_seccion(ROW,
        '5.  RELACIONES INTERPERSONALES',
        'Calidad de las relaciones en las últimas 4 semanas. '
        'Excluye respuestas "No aplica".')
    ROW = titulo_tabla(ROW, '5.1', 'Distribución por Tipo de Vínculo')

    # Detectar categorías presentes
    cats_pres = [cat for cat in CATS_REL if cat in T51.columns]
    ROW = encabezados(ROW, ['Vínculo', 'N aplica'] + cats_pres)
    for i, row_d in T51.iterrows():
        vals_fila = [row_d['N aplica']] + [row_d.get(cat, '—') for cat in cats_pres]
        ROW = fila_dato(ROW, row_d['Vínculo'], vals_fila, alt=i%2==0)
    ROW = spacer(ROW, 2)

# ══════════════════════════════════════════════════════════════════════════════
# SECCIÓN 6 — SATISFACCIÓN DE VIDA
# ══════════════════════════════════════════════════════════════════════════════
if T61 is not None and len(T61) > 0:
    ROW = titulo_seccion(ROW,
        '6.  SATISFACCIÓN DE VIDA',
        'Escala de 0 a 10 — donde 0 = nada satisfecho y 10 = muy satisfecho')
    ROW = titulo_tabla(ROW, '6.1', 'Promedio de Satisfacción por Dimensión')
    ROW = encabezados(ROW, ['Dimensión', 'Promedio (0–10)', 'Mín.', 'Máx.', 'N válido'])
    for i, row_d in T61.iterrows():
        ROW = fila_dato(ROW, row_d['Dimensión'],
                        [row_d['Promedio (0–10)'], row_d['Mínimo'],
                         row_d['Máximo'], row_d['N válido']], alt=i%2==0)
    ROW = spacer(ROW, 2)

# ── Nota al pie ───────────────────────────────────────────────────────────────
ws.row_dimensions[ROW].height = 14
ws.merge_cells(f'B{ROW}:H{ROW}')
c = ws.cell(ROW, 2,
    f'  Fuente: {os.path.basename(INPUT_FILE)}  ·  '
    f'Generado automáticamente · SCRIPT_IRT_Universal_Caracterizacion_Excel v1.1')
c.font      = Font(italic=True, size=7, color='AAAAAA', name='Arial')
c.alignment = Alignment(horizontal='left', vertical='center', indent=1)

wb.save(OUTPUT_FILE)
print(f'\n{"=" * 60}')
print(f'  ✅  LISTO  →  {OUTPUT_FILE}')
print(f'      {N_total} pacientes  ·  {PERIODO}')
print(f'{"=" * 60}')
